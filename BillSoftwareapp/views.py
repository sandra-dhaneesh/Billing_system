#Multiuserbillingindia
from django.shortcuts import render,redirect
from django.contrib.auth.models import User, auth
from . models import *
from django.contrib import messages
from django.utils.crypto import get_random_string
from django.contrib.auth.decorators import login_required
from django.utils.text import capfirst
from datetime import date
from django.template.response import TemplateResponse
from django.http.response import JsonResponse
from django.db.models import Sum
from django.shortcuts import get_object_or_404
from django.db.models import F
from openpyxl import load_workbook
from django.views.decorators.http import require_POST
from django.views.decorators.csrf import csrf_exempt
from django.db.models import Sum
from .models import SalesInvoice
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import NamedStyle
from datetime import datetime
from openpyxl import load_workbook
from django.http import JsonResponse
from django.db import transaction
import pandas as pd
from django.http import JsonResponse
from django.http import JsonResponse
from django.shortcuts import render, redirect
from openpyxl import load_workbook
from django.views import View
from django.template.loader import get_template
from django.core.mail import EmailMessage
from django.conf import settings
from io import BytesIO
# from xhtml2pdf import pisa
from django.core.mail import send_mail
from django.utils.dateparse import parse_date
# Create your views here.

def home(request):
  return render(request, 'home.html')


def about(request):
  return render(request, 'about.html')

def contact(request):
  return render(request, 'contact.html')


def service(request):
  return render(request, 'service.html')


def homepage(request):
  sid = request.session.get('staff_id')
  staff =  staff_details.objects.get(id=sid)
  cmp = company.objects.get(id=staff.company.id)
  context={
   'staff':staff,
    'cmp':cmp,
  }
  return render(request, 'companyhome.html',context)


def staffhome(request):
  sid = request.session.get('staff_id')
  staff =  staff_details.objects.get(id=sid)
  cmp = company.objects.get(id=staff.company.id)
  context={
   'staff':staff,
    'cmp':cmp,
  }
  return render(request, 'staffhome.html',context)

def register(request):
  return render(request, 'register.html')

def registercompany(request):
  return render(request, 'registercompany.html')

def registerstaff(request):
  return render(request, 'registerstaff.html')

def login(request):
  return render(request, 'login.html')

def logout(request):
    auth.logout(request)
    return redirect('/')



def registeruser(request):
    if request.method == 'POST':
        first_name = request.POST['fname']
        last_name = request.POST['lname']
        user_name = request.POST['username']
        email_id = request.POST['email']
        mobile = request.POST['phoneno']
        passw = request.POST['pass']
        c_passw = request.POST['re_pass']
        profile_pic = request.FILES.get('image')

        # Additional validation checks
        if passw != c_passw:
            messages.error(request, 'Passwords do not match')
            return redirect('register')

        if User.objects.filter(username=user_name).exists():
            messages.error(request, 'Sorry, Username already exists')
            return redirect('register')

        if User.objects.filter(email=email_id).exists():
            messages.error(request, 'Sorry, Email already exists')
            return redirect('register')

        # if not email_id.endswith('@gmail.com'):
        #     messages.error(request, 'Invalid email address')
        #     return redirect('register')

        user_data = User.objects.create_user(
            first_name=first_name,
            last_name=last_name,
            username=user_name,
            email=email_id,
            password=passw
        )
        user_data.save()

        data = User.objects.get(id=user_data.id)
        cust_data = company(contact=mobile, profile_pic=profile_pic, user=data)
        cust_data.save()

        demo_staff = staff_details(
            company=cust_data,
            email=email_id,
            position='company',
            user_name=user_name,
            password=passw,
            contact=mobile
        )
        demo_staff.save()

        # messages.success(request, 'Registration successful')
        return redirect('registercompany')

    return render(request, 'register.html')


def add_company(request):
    if request.method == 'POST':
        email = request.POST['email']
        
        try:
            user = User.objects.get(email=email)
        except User.DoesNotExist:
            messages.error(request, 'Error: User with the specified email does not exist.')
            return render(request, 'registercompany.html')

        c = company.objects.get(user=user)
        c.company_name = request.POST['companyname']
        if company.objects.filter(company_name=c.company_name).exists():
            messages.error(request, 'Error: Company with the specified name already exists.')
            return render(request, 'registercompany.html')
        c.address = request.POST['address']
        c.city = request.POST['city']
        c.state = request.POST['state']
        c.country = request.POST['country']
        c.pincode = request.POST['pincode']
        c.pan_number = request.POST['pannumber']
        c.gst_type = request.POST['gsttype']
        c.gst_no = request.POST['gstno']

        code = get_random_string(length=6)
        if company.objects.filter(Company_code=code).exists():
            code2 = get_random_string(length=6)
            c.Company_code = code2
        else:
            c.Company_code = code
        
        c.save()

        staff = staff_details.objects.get(email=email, position='company', company=c)
        staff.first_name = request.POST['companyname']
        staff.last_name = ''
        staff.save()

        return redirect('login')  

    return render(request, 'registercompany.html')


def staff_registraction(request):
  if request.method == 'POST':
    fn=request.POST['fname']
    ln=request.POST['lname']
    email=request.POST['email']
    un=request.POST['username']
    ph=request.POST['phoneno']
    pas=request.POST['pass']
    code=request.POST['companycode']

    if company.objects.filter(Company_code=code).exists():
      com=company.objects.get(Company_code=code)
    else:
        messages.info(request, 'Sorry, Company code is Invalid')
        return redirect('registerstaff')
    img=request.FILES.get('image')

    if staff_details.objects.filter(user_name=un,company=com).exists():
      messages.info(request, 'Sorry, Username already exists')
      return redirect('registerstaff')
    elif staff_details.objects.filter(email=email,company=com).exists():
      messages.info(request, 'Sorry, Email already exists')
      return redirect('registerstaff')
    elif staff_details.objects.filter(contact=ph,company=com).exists():
      messages.info(request, 'Sorry, Phone Number already exists')
      return redirect('registerstaff')
    elif staff_details.objects.filter(password=pas).exists():
      messages.info(request, 'Sorry, Password already exists,please use another password')
      return redirect('registerstaff')
    else:
      
      staff=staff_details(first_name=fn,last_name=ln,email=email,user_name=un,contact=ph,password=pas,img=img,company=com)
      staff.save()
      return redirect('login')

  else:
    print(" error")
    return redirect('registerstaff')
  

  
def loginurl(request):
  if request.method == 'POST':
    user_name = request.POST['username']
    passw = request.POST['pass']
    
    
    log_user = auth.authenticate(username = user_name,
                                  password = passw)
    
    if log_user is not None:
      auth.login(request, log_user)
        
    if staff_details.objects.filter(user_name=user_name,password=passw,position='company').exists():
      data = staff_details.objects.get(user_name=user_name,password=passw,position='company') 

      request.session["staff_id"]=data.id
      if 'staff_id' in request.session:
        if request.session.has_key('staff_id'):
          staff_id = request.session['staff_id']
          print(staff_id)
 
        return redirect('homepage')  

    if staff_details.objects.filter(user_name=user_name,password=passw,position='staff').exists():
      data = staff_details.objects.get(user_name=user_name,password=passw,position='staff')   

      request.session["staff_id"]=data.id
      if 'staff_id' in request.session:
        if request.session.has_key('staff_id'):
          staff_id = request.session['staff_id']
          print(staff_id)
 
          return redirect('staffhome')  
    else:
      messages.info(request, 'Invalid Username or Password. Try Again.')
      return redirect('login')  
  else:  
   return redirect('login')   
  
  

@login_required(login_url='login')  
def profile(request):
    if request.user.is_authenticated:
        try:
            com = company.objects.get(user=request.user)
            staff_id = request.session['staff_id']
            staff =  staff_details.objects.get(id=staff_id)
            context = {
               'company': com,
               'staff':staff
               }
            return render(request, 'profile.html', context)
        except company.DoesNotExist:
            messages.error(request, 'Company not found for the authenticated user.')
            return redirect('login')
    else:
        messages.info(request, 'Please log in to view your profile.')
        return redirect('login')

    

def editprofile(request,pk):
  com= company.objects.get(id = pk)
  context={
     'company':com
  }
  return render(request, 'editprofile.html',context)

def edit_profilesave(request,pk):
  com= company.objects.get(id = pk)
  user1 = User.objects.get(id = com.user_id)

  if request.method == "POST":

      user1.first_name = capfirst(request.POST.get('f_name'))
      user1.last_name  = capfirst(request.POST.get('l_name'))
      user1.email = request.POST.get('email')
      com.contact = request.POST.get('cnum')
      com.address = capfirst(request.POST.get('ards'))
      com.company_name = request.POST.get('comp_name')
      # user1.email = request.POST.get('comp_email')
      com.city = request.POST.get('city')
      com.state = request.POST.get('state')
      com.country = request.POST.get('country')
      com.pincode = request.POST.get('pinc')
      com.gst_type = request.POST.get('gsttype')
      com.gst_no = request.POST.get('gstno')
      com.pan_number = request.POST.get('pan')
      if len(request.FILES)!=0 :
          com.profile_pic = request.FILES.get('file')
          

      com.save()
      user1.save()
      return redirect('profile')

  context = {
      'company' : com,
      'user1' : user1,
  } 

  return render(request,'editprofile.html',context)

def base(request):
  staff_id = request.session['staff_id']
  staff =  staff_details.objects.get(id=staff_id)
  context = {
              'staff' : staff

          }
  return render(request, 'base.html',context)

def staffprofile(request):
  staff_id = request.session['staff_id']
  staff =  staff_details.objects.get(id=staff_id)
  context = {
              'staff' : staff

          }
  return render(request,'profilestaff.html',context)

def editstaffprofile(request):
  staff_id = request.session['staff_id']
  staff =  staff_details.objects.get(id=staff_id)
  context={
     'staff':staff
  }
  return render(request, 'editstaff.html',context)

def edit_staffprofilesave(request):
  staff_id = request.session['staff_id']
  staff =  staff_details.objects.get(id=staff_id)

  if request.method == "POST":

      staff.first_name = capfirst(request.POST.get('f_name'))
      staff.last_name  = capfirst(request.POST.get('l_name'))
      staff.email = request.POST.get('email')
      staff.contact = request.POST.get('cnum')
      if len(request.FILES)!=0 :
          staff.img = request.FILES.get('file')

      staff.save()
      return redirect('staffprofile')

  context = {
      'staff' : staff
  } 

  return render(request,'editstaff.html',context)

def add_item(request):
  sid = request.session.get('staff_id')
  staff =  staff_details.objects.get(id=sid)
  cmp = company.objects.get(id=staff.company.id)
  todate = date.today()
  tdate = todate.strftime("%Y-%m-%d")
  item_units = ItemUnitModel.objects.filter(company = cmp)
  context={
    'staff':staff,
    'cmp':cmp,
    'date':tdate,
    'item_units':item_units
  }
  return render(request, 'add_item.html',context)


def item_create_new(request):
    if request.method == 'POST':
        sid = request.session.get('staff_id')
        staff = staff_details.objects.get(id=sid)
        cmp = company.objects.get(id=staff.company.id)

        item_name = request.POST.get('item_name')
        item_hsn = request.POST.get('item_hsn')

        if ItemModel.objects.filter(item_name=item_name, company=cmp).exists() or \
                ItemModel.objects.filter(item_hsn=item_hsn, company=cmp).exists():
            messages.error(request, 'Item Name or HSN already exists. Please choose a different Name or HSN.')
            return render(request, 'add_item.html')
        if len(item_hsn) < 6:
            messages.error(request, 'Item HSN must be 6 digits or more.')
            return render(request, 'add_item.html')

        item_unit = request.POST.get('item_unit')
        item_type = request.POST.get('type')
        item_taxable = request.POST.get('item_taxable')
        item_gst = request.POST.get('item_gst')
        item_igst = request.POST.get('item_igst')
        item_sale_price = request.POST.get('saleprice')
        item_purchase_price = request.POST.get('purprice')
        item_opening_stock = request.POST.get('item_opening_stock')
        item_current_stock = item_opening_stock
        if item_opening_stock == '' or None:
            item_opening_stock = 0
            item_current_stock = 0
        item_at_price = request.POST.get('item_at_price')
        if item_at_price == '' or None:
            item_at_price = 0
        item_date = request.POST.get('itmdate')

        item_data = ItemModel(user=staff.company.user,
                              company=cmp,
                              staff=staff,
                              item_name=item_name,
                              item_hsn=item_hsn,
                              item_unit=item_unit,
                              item_type=item_type,
                              item_taxable=item_taxable,
                              item_gst=item_gst,
                              item_igst=item_igst,
                              item_sale_price=item_sale_price,
                              item_purchase_price=item_purchase_price,
                              item_stock_in_hand=item_opening_stock,
                              item_current_stock=item_current_stock,
                              item_at_price=item_at_price,
                              item_date=item_date)
        item_data.save()

        tr_history = ItemTransactionHistory(company=cmp,
                                            staff=staff,
                                            item=item_data,
                                            action="CREATED",
                                            done_by_name=staff.first_name,)
        tr_history.save()

        if request.POST.get('save_and_next'):
            return redirect('add_item')
        elif request.POST.get('itemsave'):
            return redirect('view_item')

    return render(request, 'add_item.html')



def view_item(request):
  sid = request.session.get('staff_id')
  staff =  staff_details.objects.get(id=sid)
  cmp = company.objects.get(id=staff.company.id)
  allitem = ItemModel.objects.filter(company=cmp)
  # for i in allitem:
  #     last_transaction = ItemTransactionHistory.objects.filter(item=i).last()
  #     if last_transaction:
  #         i.action = last_transaction.action
  #         i.done_by_name = last_transaction.done_by_name
  #     else:
  #         i.action = None
  #         i.done_by_name = None

  context={
    'staff':staff,
    'cmp':cmp,
    'allitem':allitem,
  }
  return render(request, 'view_item.html',context)

def view_items(request, pk):
    sid = request.session.get('staff_id')
    staff = staff_details.objects.get(id=sid)
    cmp = company.objects.get(id=staff.company.id)
    allitem = ItemModel.objects.filter(company=cmp)
    items = ItemModel.objects.get(id=pk)
    if pk == 0:
      first_item = allitem.filter().first()
    else:
      first_item = allitem.get(id=pk)
      transactions = ItemTransactionModel.objects.filter(company = cmp,item=first_item.id).order_by('-trans_created_date')
      

    # last_transaction = ItemTransactionHistory.objects.filter(item=items).last()
    # if last_transaction:
    #     items.action = last_transaction.action
    #     items.done_by_name = last_transaction.done_by_name
    # else:
    #     items.action = None
    #     items.done_by_name = None

    context = {
        'staff': staff,
        'cmp': cmp,
        'item': items,
        'first_item':first_item,
        'allitem':allitem,
        'transactions':transactions,
        'item_name': items.item_name,
    }

    return render(request, 'view_items.html', context)


def edit_item(request,pk):
  sid = request.session.get('staff_id')
  staff = staff_details.objects.get(id=sid)
  cmp = company.objects.get(id=staff.company.id)
  allitem = ItemModel.objects.filter(company=cmp)
  items = ItemModel.objects.get(id=pk)

  context = {
        'staff': staff,
        'cmp': cmp,
        'item': items,
        'allitem':allitem
    }

  return render(request, 'edit_item.html',context)


# def update_item(request,pk):
#   if request.method=='POST':
#     sid = request.session.get('staff_id')
#     staff =  staff_details.objects.get(id=sid)
#     cmp = company.objects.get(id=staff.company.id)
#     item = ItemModel.objects.get(id=pk)

#     item.item_name = request.POST.get('item_name')
#     item.item_hsn = request.POST.get('item_hsn')
#     item.item_unit = request.POST.get('item_unit')
#     item.item_type = request.POST.get('type')
#     item.item_taxable = request.POST.get('item_taxable')
#     item.item_gst = request.POST.get('item_gst')
#     item.item_igst = request.POST.get('item_igst')
#     item.item_sale_price = request.POST.get('saleprice')
#     item.item_purchase_price = request.POST.get('purprice')
#     item.item_stock_in_hand = request.POST.get('item_opening_stock')
#     item.item_current_stock =item.item_stock_in_hand
#     item.item_at_price = request.POST.get('item_at_price')
#     item.item_date = request.POST.get('itmdate')
    
#     item.save()

#     tr_history = ItemTransactionHistory(company=cmp,
#                                             staff=staff,      
#                                             item=item,
#                                             action="UPDATED",
#                                             done_by_name=staff.first_name,
#                                             )
#     tr_history.save()
  

#     return redirect('view_item')
#   return redirect('edit_item')

def update_item(request, pk):
    if request.method == 'POST':
        sid = request.session.get('staff_id')
        staff = staff_details.objects.get(id=sid)
        cmp = company.objects.get(id=staff.company.id)

        item_data = ItemModel.objects.get(id=pk)
        user = cmp.user
        company_user_data = cmp

        item_name = request.POST.get('item_name')
        item_hsn = request.POST.get('item_hsn')
        item_unit = request.POST.get('item_unit')
        item_type = request.POST.get('type')
        item_taxable = request.POST.get('item_taxable')
        item_gst = request.POST.get('item_gst')
        item_igst = request.POST.get('item_igst')
        
        if item_taxable == 'Non Taxable':
            item_gst = 'GST0[0%]'
            item_igst = 'IGST0[0%]'

        item_sale_price = request.POST.get('saleprice')
        item_purchase_price = request.POST.get('purprice')
        item_stock_in_hand = request.POST.get('item_opening_stock')

        if item_stock_in_hand == '':
            item_stock_in_hand = 0

        item_current_stock = item_data.item_current_stock

        if int(item_stock_in_hand) > item_data.item_stock_in_hand:
            item_stock = item_data.item_current_stock + (int(item_stock_in_hand) - item_data.item_stock_in_hand)
        elif int(item_stock_in_hand) < item_data.item_stock_in_hand:
            item_stock = item_data.item_current_stock - (item_data.item_stock_in_hand - int(item_stock_in_hand))
        else:
            item_stock = item_current_stock

        item_at_price = request.POST.get('item_at_price')

        if item_at_price == '':
            item_at_price = 0

        item_date = request.POST.get('itmdate')

        item_data.user = user
        item_data.company = company_user_data
        item_data.item_name = item_name
        item_data.item_hsn = item_hsn
        item_data.item_unit = item_unit
        item_data.item_type = item_type
        item_data.item_taxable = item_taxable
        item_data.item_gst = item_gst
        item_data.item_igst = item_igst
        item_data.item_sale_price = item_sale_price
        item_data.item_purchase_price = item_purchase_price
        item_data.item_stock_in_hand = item_stock_in_hand
        item_data.item_current_stock = int(item_stock)
        item_data.item_at_price = item_at_price
        item_data.item_date = item_date

        item_data.save()
        print('\nupdated')

        tr_history = ItemTransactionHistory(company=cmp,
                                        staff=staff,      
                                        item=item_data,
                                        action="UPDATED",
                                        done_by_name=staff.first_name,
                                        )
        tr_history.save()

    return redirect('view_item')



def item_delete(request,pk):
  item_to_delete = ItemModel.objects.get(id=pk)
  item_to_delete.delete()
  return redirect('view_item')

def itemhistory(request,pk):
  staff_id = request.session['staff_id']
  staff =  staff_details.objects.get(id=staff_id)
  cmp = company.objects.get(id=staff.company.id)
  history = ItemTransactionHistory.objects.filter(item=pk)

  context = {
              'staff' : staff,
              'history':history,

          }
  return render(request,'itemtranstationhistory.html',context)

def itemmodaladjust(request,pk):
  item = ItemModel.objects.get(id=pk)
  return TemplateResponse(request,'itemmodaladjust.html',{"item":item,})


def ajust_quantity(request,pk):
  sid = request.session.get('staff_id')
  staff =  staff_details.objects.get(id=sid)
  cmp = company.objects.get(id=staff.company.id)

  if request.method=='POST':
    item = ItemModel.objects.get(id=pk)

    # user = User.objects.get(id=request.user.id)
    user = cmp.user
    # company_user_data = company.objects.get(user=request.user.id)
    company_user_data = cmp

    trans_type_check_checked = request.POST.get('trans_type')
    if trans_type_check_checked == 'on':
      trans_type = 'reduce stock'
      trans_qty = request.POST.get('reduced_qty')
    else:
      trans_type = 'add stock'
      trans_qty = request.POST.get('added_qty')
    trans_user_name = user.first_name
    trans_date = request.POST.get('trans_date')

    trans_adjusted_qty= request.POST.get('adjusted_qty')
    trans_current_qty = request.POST.get('item_qty')
    print(f'the quantity : {trans_current_qty}')
    item.item_current_stock = trans_adjusted_qty
    item.save()
    transaction_data = ItemTransactionModel(user=user,
                                        company=company_user_data,
                                        staff=staff,
                                        item=item,
                                        trans_type=trans_type,
                                        trans_user_name=trans_user_name,
                                        trans_date=trans_date,
                                        trans_qty=trans_qty,
                                        trans_current_qty=trans_current_qty,
                                        trans_adjusted_qty=trans_adjusted_qty,)
    transaction_data.save()
  return redirect('view_items',pk=item.id)

def item_unit_create(request):
  if request.method=='POST':
    #updated-shemeem
    sid = request.session.get('staff_id')
    staff =  staff_details.objects.get(id=sid)
    cmp = company.objects.get(id=staff.company.id)

    # user = User.objects.get(id=request.user.id)
    # company_user_data = company.objects.get(user=request.user.id)

    item_unit_name = request.POST.get('item_unit_name')
    unit_data = ItemUnitModel(user=cmp.user,company=cmp,unit_name=item_unit_name)
    unit_data.save()
  return redirect('add_item')


def edititemmodaladjust(request,pk,trans):
  item = ItemModel.objects.get(id=pk)
  transaction = ItemTransactionModel.objects.get(id=trans)
  print('enterd')
  return render(request,'edititemmodaladjust.html',{"item":item,"transaction":transaction,})

def update_adjusted_transaction(request,pk,trans):
  item = ItemModel.objects.get(id=pk)
  transaction = ItemTransactionModel.objects.get(id=trans)
  sid = request.session.get('staff_id')
  staff =  staff_details.objects.get(id=sid)
  cmp = company.objects.get(id=staff.company.id)

  if request.method=='POST':
    item = ItemModel.objects.get(id=pk)

    user = cmp.user
    company_user_data = cmp

    trans_type_check_checked = request.POST.get('trans_type')
    if trans_type_check_checked == 'on':
      trans_type = 'reduce stock'
      trans_qty = request.POST.get('reduced_qty')
    else:
      trans_type = 'add stock'
      trans_qty = request.POST.get('added_qty')
    trans_user_name = user.first_name
    trans_date = request.POST.get('trans_date')

    adjusted_qty= request.POST.get('adjusted_qty')
    trans_current_qty = request.POST.get('item_qty')
    if transaction.trans_type == 'reduce stock':
      if trans_type == 'reduce stock':
        print('reduce to reduce')
        item.item_current_stock = item.item_current_stock - (int(trans_qty)  - transaction.trans_qty)
      else:
        print('reduce to add')
        print(f'{trans_qty}-{transaction.trans_qty}={((int(trans_qty)  - transaction.trans_qty))}')
        item.item_current_stock = item.item_current_stock + transaction.trans_qty + int(trans_qty)
    else:
      if trans_type == 'reduce stock':
        print('add to red')
        item.item_current_stock = item.item_current_stock - (int(trans_qty)  + transaction.trans_qty)
      else:
        print('add to add')
        print(f'{trans_qty}-{transaction.trans_qty}={((int(trans_qty)  - transaction.trans_qty))}')
        item.item_current_stock = item.item_current_stock - transaction.trans_qty + int(trans_qty)
    # item.item_opening_stock = adjusted_qty
    item.save()
    transaction.trans_type =trans_type
    transaction.trans_date=trans_date
    transaction.trans_qty =trans_qty
    transaction.trans_current_qty=trans_current_qty
    transaction.save()
  return redirect('view_items',pk=item.id)

def transaction_delete(request,pk):
  transaction = ItemTransactionModel.objects.get(id=pk)
  item = ItemModel.objects.get(id=transaction.item_id)
  print(transaction.trans_type)
  if transaction.trans_type=='add stock':
    print('add')
    item.item_current_stock = item.item_current_stock - transaction.trans_qty
    print(item.item_name)
    print(item.item_current_stock)
    print(item.item_current_stock)
    print(transaction.trans_qty)
    print(item.item_current_stock - transaction.trans_qty)
  else:
    print('reduce')
    item.item_current_stock = item.item_current_stock + transaction.trans_qty
  item.save()
  transaction.delete()
  return redirect('view_items',pk=item.id)

def item_delete_openstock(request,pk):
  item = ItemModel.objects.get(id=pk)
  if item.item_stock_in_hand > item.item_current_stock:
    item.item_current_stock =item.item_stock_in_hand - item.item_current_stock
  else:
    item.item_current_stock =item.item_current_stock - item.item_stock_in_hand
  # item.item_current_stock =  item.item_opening_stock - item.item_current_stock
  item.item_stock_in_hand = 0
  # print(f'{item.item_current_stock }={item.item_opening_stock}-{item.item_current_stock}')
  item.save()
  return redirect('view_items',pk=item.id)
  


def sharedebitToEmail(request,id):
  if request.user:
        try:
            if request.method == 'POST':
                emails_string = request.POST['email_ids']

                # Split the string by commas and remove any leading or trailing whitespace
                emails_list = [email.strip() for email in emails_string.split(',')]
                email_message = request.POST['email_message']
                print(emails_list)

                sid = request.session.get('staff_id')
                staff =  staff_details.objects.get(id=sid)
                cmp = company.objects.get(id=staff.company.id) 
               
                pdebt = purchasedebit.objects.get(pdebitid=id,company=cmp)
                pitm = purchasedebit1.objects.filter(pdebit=pdebt,company=cmp)
                        
                context = {'pdebt':pdebt, 'cmp':cmp,'pitm':pitm}
                template_path ='debitnote_file_mail.html'
                template = get_template(template_path)

                html  = template.render(context)
                result = BytesIO()
                pdf = pisa.pisaDocument(BytesIO(html.encode("ISO-8859-1")), result)
                pdf = result.getvalue()
                filename = f'DEBIT NOTE - {pdebt.pdebitid}.pdf'
                subject = f"DEBIT NOTE - {pdebt.pdebitid}"
                email = EmailMessage(subject, f"Hi,\nPlease find the attached DEBIT NOTE - File-{pdebt.pdebitid}. \n{email_message}\n\n--\nRegards,\n{cmp.company_name}\n{cmp.address}\n{cmp.state} - {cmp.country}\n{cmp.contact}", from_email=settings.EMAIL_HOST_USER, to=emails_list)
                email.attach(filename, pdf, "application/pdf")
                email.send(fail_silently=False)

                msg = messages.success(request, 'Debit note file has been shared via email successfully..!')
                return redirect(details_debitnote,id)
        except Exception as e:
            print(e)
            messages.error(request, f'{e}')
            return redirect(details_debitnote, id)



def delete_debit(request,id):
  sid = request.session.get('staff_id')
  staff = staff_details.objects.get(id=sid)
  cmp = company.objects.get(id=staff.company.id) 
  pdebt = purchasedebit.objects.get(pdebitid=id)
  purchasedebit1.objects.filter(pdebit=pdebt,company=cmp).delete()
  pdebt.delete()
  return redirect('view_purchasedebit')


def details_debitnote(request,id):
  sid = request.session.get('staff_id')
  staff =  staff_details.objects.get(id=sid)
  cmp = company.objects.get(id=staff.company.id) 
  allmodules = modules_list.objects.filter(company=staff.company,status='New')
  pdebt = purchasedebit.objects.get(pdebitid=id,company=cmp)
  pitm = purchasedebit1.objects.filter(pdebit=pdebt,company=cmp)
  dis = 0
  for itm in pitm:
    dis += int(itm.discount)
  itm_len = len(pitm)

  context={'staff':staff,'allmodules':allmodules,'pdebt':pdebt,'pitm':pitm,'itm_len':itm_len,'dis':dis}
  return render(request,'debitnotedetails.html',context)


def edit_debitnote(request,id):
  toda = date.today()
  tod = toda.strftime("%Y-%m-%d")
  sid = request.session.get('staff_id')
  staff =  staff_details.objects.get(id=sid)
  cmp = company.objects.get(id=staff.company.id)
  party = Parties.objects.filter(company=cmp)
  item = ItemModel.objects.filter(company=cmp,user=cmp.user)
  item_units = ItemUnitModel.objects.filter(user=cmp.user,company=staff.company.id)
  allmodules= modules_list.objects.filter(company=staff.company,status='New')
  pdebt = purchasedebit.objects.get(pdebitid=id,company=cmp)
  debtitem = purchasedebit1.objects.filter(pdebit=pdebt,company=cmp)

  

  billno=PurchaseBill.objects.filter(company=cmp,party=pdebt.party).values('billno')
  billdate=PurchaseBill.objects.filter(company=cmp,party=pdebt.party).values('billdate')
  print(billno)
  print(billdate)
  ddate = pdebt.debitdate.strftime("%Y-%m-%d")

 



  context = {'staff':staff,  'allmodules':allmodules, 'pdebt':pdebt, 'debtitem':debtitem, 'party':party, 'item':item, 'item_units':item_units, 'ddate':ddate,'tod':tod,'billno':billno,'billdate':billdate}
  return render(request,'debitnoteedit.html',context)



def update_debitnote(request,id):
  if request.method =='POST':
    sid = request.session.get('staff_id')
    staff = staff_details.objects.get(id=sid)
    cmp = company.objects.get(id=staff.company.id)  
    partys = Parties.objects.get(id=request.POST.get('customername'))
    pdebt = purchasedebit.objects.get(pdebitid=id,company=cmp)
    pdebt.party = partys
    pdebt.debitdate = request.POST.get('debitdate')
    pdebt.billno = request.POST.get('billno')
    pdebt.billdate = request.POST.get('billdate')
    pdebt.supply  = request.POST.get('placosupply')
    pdebt.subtotal =float(request.POST.get('subtotal'))
    pdebt.grandtotal = request.POST.get('grandtotal')
    pdebt.igst = request.POST.get('igst')
    pdebt.cgst = request.POST.get('cgst')
    pdebt.sgst = request.POST.get('sgst')
    pdebt.taxamount = request.POST.get("taxamount")
    pdebt.adjustment = request.POST.get("adj")
 


    pdebt.save()

    product = tuple(request.POST.getlist("product[]"))
    qty = tuple(request.POST.getlist("qty[]"))
    total = tuple(request.POST.getlist("total[]"))
    discount = tuple(request.POST.getlist("discount[]"))
    tax =  tuple(request.POST.getlist("tax[]"))
    itemId=request.POST.getlist("itemId[]")

    purchasedebit1.objects.filter(pdebit=pdebt,company=cmp).delete()
    if len(total)==len(discount)==len(qty)==len(itemId)==len(tax):
      mapped=zip(product,qty,discount,total,itemId,tax)
      mapped=list(mapped)
      for ele in mapped:
        itm = ItemModel.objects.get(id=ele[0])
        purchasedebit1.objects.create(product =itm,qty=ele[1],discount=ele[2],total=ele[3],tax=ele[5],pdebit=pdebt,company=cmp)

    DebitnoteTransactionHistory.objects.create(debitnote=pdebt,company=cmp,staff=staff,action='Updated')
    return redirect('view_purchasedebit')

  return redirect('view_purchasedebit')


# def add_debitnote(request):
#   toda = date.today()
#   tod = toda.strftime("%Y-%m-%d")

#   sid = request.session.get('staff_id')
#   staff =  staff_details.objects.get(id=sid)
#   cmp = company.objects.get(id=staff.company.id)
#   cust = Parties.objects.filter(company=cmp)
#   pdebt= purchasedebit.objects.filter(company=cmp)

  
  
  
#   debt_count = purchasedebit.objects.filter(company=cmp).order_by('-pdebitid').first()
  
#   if debt_count:
#     next_count = int(debt_count.reference_number) + 1
#   else:
#     next_count=1

#   item = ItemModel.objects.filter(company=cmp,user=cmp.user)

#   context = {'staff':staff, 'cust':cust, 'cmp':cmp,'count':next_count, 'tod':tod, 'item':item, }
#   return render(request,'adddebitnotes.html',context)

def credit_save(request):
    if request.method == 'POST':
        sid = request.session.get('staff_id')
        staff = staff_details.objects.get(id=sid)
        cmp = company.objects.get(id=staff.company.id)

        # Retrieve party details if available
        party_id = request.POST.get('partyname')
        party = None
        if party_id:
            party = Parties.objects.get(id=party_id)

        # Check if 'billNo' exists in POST data and assign value accordingly
        bill_no = request.POST.get('billNo', None)
        if bill_no == 'nobill':
            bill_no = None

        # Check if 'billDate' exists in POST data and assign value accordingly
        bill_date_str = request.POST.get('billDate', None)
        bill_date = None
        if bill_date_str:
            bill_date = parse_date(bill_date_str)

        # Create an instance of Creditnote model and save the data
        credit_note = Creditnote(
            party_name=party.party_name if party else None,
            contact=party.phone_number if party else None,
            address=party.billing_address if party else None,
            invoice_no=bill_no,
            idate=bill_date,
            state_of_supply=request.POST.get('placosupply'),
            date=request.POST.get('date'),
            gstin=request.POST.get('gstin'),
            subtotal=request.POST.get('subtotal'),
            sgst=request.POST.get('sgst'),
            cgst=request.POST.get('cgst'),
            igst=request.POST.get('igst'),
            taxamount=request.POST.get('taxamount'),
            roundoff=request.POST.get('adj'),
            grandtotal=request.POST.get('grandtotal'),
            description=request.POST.get('des'),
            returnno=request.POST.get('returnno'),
            staff=staff,
            company=cmp,
            party=party
        )

        # Save the instance
        credit_note.save()
        history = CreditnoteHistory(company=cmp, staff=staff, credit=credit_note, action='Created')
        history.save()

        # Save credit note items
        product = request.POST.getlist("product[]")
        qty = request.POST.getlist("qty[]")
        discount = request.POST.getlist("discount[]")
        total = request.POST.getlist("total[]")
        hsn = request.POST.getlist("hsn[]")
        tax = request.POST.getlist("tax[]")
        price = request.POST.getlist("price[]")

        if len(product) == len(qty) == len(discount) == len(total) == len(hsn) == len(tax) == len(price):
            mapped = zip(product, qty, discount, total, hsn, tax, price)
            for ele in mapped:
                itm = ItemModel.objects.get(id=ele[0])
                CreditnoteItem.objects.create(
                    product=itm.item_name,
                    qty=ele[1],
                    discount=ele[2],
                    total=ele[3],
                    hsn=ele[4],
                    tax=ele[5],
                    price=ele[6],
                    company=cmp,
                    credit=credit_note,
                    staff=staff
                )

        Creditnote.objects.filter(company=cmp, staff=staff).update(returnno=F('returnno'))

        if 'Next' in request.POST:
            return redirect('transactiontable')

        if "Save" in request.POST:
            return redirect('credit_add')

    else:
        return render(request, 'credit_add.html')



def add_debitnote(request):
  toda = date.today()
  tod = toda.strftime("%Y-%m-%d")
  if 'staff_id' in request.session:
    if request.session.has_key('staff_id'):
      staff_id = request.session['staff_id']
           
    else:
      return redirect('/')
  staff =  staff_details.objects.get(id=staff_id)
  print("hii")
  print(staff)
  cmp = company.objects.get(id=staff.company.id)
  party = Parties.objects.filter(company=cmp)

  allmodules = modules_list.objects.filter(company=staff.company, status='New').first()
  item=ItemModel.objects.filter(company=cmp,user=cmp.user)
  item_units = ItemUnitModel.objects.filter(user=cmp.user,company=staff.company)
  billno=PurchaseBill.objects.filter(company=cmp).values('billno')
  
 
  debt_count = purchasedebit.objects.filter(company=cmp).order_by('-pdebitid').first()
  
  if debt_count:
    next_count = int(debt_count.reference_number) + 1
  else:
    next_count = 1

  print(billno)
  return render(request,'adddebitnotes.html',{'staff':staff,'allmodules':allmodules,'party':party,'item':item,'count':next_count,'item_units':item_units, 'tod':tod, 'cmp':cmp,'billno':billno})


def add_parties(request):
  
  sid = request.session.get('staff_id')
  staff =  staff_details.objects.get(id=sid)
  cmp = company.objects.get(id=staff.company.id)

  return render(request, 'add_parties.html',{'staff':staff})

def save_parties(request):
    if request.method == 'POST':
        staff_id = request.session['staff_id']
        staff =  staff_details.objects.get(id=staff_id)
        
        party_name = request.POST['partyname']
        gst_no = request.POST['gstno']
        contact = request.POST['contact']
        gst_type = request.POST['gst']
        state = request.POST['state']
        address = request.POST['address']
        email = request.POST['email']
        openingbalance = request.POST.get('balance', '')
        payment = request.POST.get('paymentType', '')
        creditlimit = request.POST.get('creditlimit', '')
        current_date = request.POST['currentdate']
        End_date = request.POST.get('enddate', None)
        additionalfield1 = request.POST['additionalfield1']
        additionalfield2 = request.POST['additionalfield2']
        additionalfield3 = request.POST['additionalfield3']
       
        if (
          not party_name
          
      ):
          return render(request, 'add_parties.html')

        part = party(party_name=party_name, gst_no=gst_no,contact=contact,gst_type=gst_type, state=state,address=address, email=email, openingbalance=openingbalance,payment=payment,
                       creditlimit=creditlimit,current_date=current_date,End_date=End_date,additionalfield1=additionalfield1,additionalfield2=additionalfield2,additionalfield3=additionalfield3,user=staff.company.user,company=staff.company)
        part.save() 

        if 'save_and_new' in request.POST:
            
            return render(request, 'add_parties.html')
        else:
          
            return redirect('view_parties')

    return render(request, 'add_parties.html')


# def create_debitnotes(request):
#     if request.method == 'POST':
#         sid = request.session.get('staff_id')
#         staff = staff_details.objects.get(id=sid)
#         cmp = company.objects.get(id=staff.company.id)
#         part = Parties.objects.get(id=request.POST.get('customername'))
#         pdebt = purchasedebit(party=part,
#                   pdebitid=request.POST.get('pdebitid'),
                      
#                       supply=request.POST.get('placosupply'),
#                       billno=request.POST.get("bill_no"),
#                       billdate=request.POST.get("billdate"), 
#                       reference_number=request.POST.get("pdebitid"),
#                       subtotal=float(request.POST.get('subtotal')),
#                       igst = request.POST.get('igst'),
#                       cgst = request.POST.get('cgst'),
#                       sgst = request.POST.get('sgst'),
#                       adjustment = request.POST.get("adj"),
#                       taxamount = request.POST.get("taxamount"),
#                       grandtotal=request.POST.get('grandtotal'),
#                       company=cmp,staff=staff)
#         pdebt.save()

#         product = tuple(request.POST.getlist("product[]"))
#         qty = tuple(request.POST.getlist("qty[]"))
#         discount = tuple(request.POST.getlist("discount[]"))
#         total = tuple(request.POST.getlist("total[]"))

#         pdebitid = purchasedebit.objects.get(pdebitid=pdebt.pdebitid, company=cmp)

#         if len(product) == len(qty) == len(discount) == len(total):
#             mapped = zip(product, qty, discount, total)
#             mapped = list(mapped)
#             for ele in mapped:
#                 itm = ItemModel.objects.get(id=ele[0])
#                 purchasedebit1.objects.create(product =itm,qty=ele[1],discount=ele[2],total=ele[3],pdebit=pdebitid,company=cmp)
       

#         purchasedebit.objects.filter(company=cmp).update(tot_debt_no=F('tot_debt_no') + 1)

#         pdebt.tot_debt_no = pdebt.pdebitid
#         pdebt.save()

#         DebitnoteTransactionHistory.objects.create(debitnote=pdebt, company=cmp, staff=staff, action='Created')

#         if 'Next' in request.POST:
#             return redirect('add_debitnote')

#         if "Save" in request.POST:
#             return redirect('view_purchasedebit')

#     else:
#         return render(request, 'adddebitnote.html')

def create_debitnotes(request):
  if 'staff_id' in request.session:
    if request.session.has_key('staff_id'):
      staff_id = request.session['staff_id']
           
    else:
      return redirect('/')
  staff =  staff_details.objects.get(id=staff_id)
  cmp = company.objects.get(id=staff.company.id)
  partys=Parties.objects.get(id=request.POST.get('customername'))
  billno = purchasedebit.objects.filter(party=partys).values_list('billno', flat=True)
  billdate = purchasedebit.objects.filter(party=partys).values_list('billdate', flat=True)

  context = {
        'billno': billno,
        'billdate':billdate
        # Add other context variables as needed
    }
  if request.method == 'POST': 
    print(request.POST.get("bill_no"),"ammu")
    pdebt = purchasedebit(party=partys,
                      pdebitid=request.POST.get('pdebitid'),
                      debitdate=request.POST.get('debitdate'),
                      supply=request.POST.get('placosupply'),
                      billno=request.POST.get("billno"),
                      billdate=request.POST.get("billdate"), 
                      reference_number=request.POST.get("pdebitid"),
                      subtotal=float(request.POST.get('subtotal')),
                      igst = request.POST.get('igst'),
                      cgst = request.POST.get('cgst'),
                      sgst = request.POST.get('sgst'),
                      adjustment = request.POST.get("adj"),
                      taxamount = request.POST.get("taxamount"),
                      grandtotal=request.POST.get('grandtotal'),
                      company=cmp,staff=staff)
    pdebt.save()

   
          
    product = tuple(request.POST.getlist("product[]"))
    qty =  tuple(request.POST.getlist("qty[]"))
    discount =  tuple(request.POST.getlist("discount[]"))
    total =  tuple(request.POST.getlist("total[]"))
    tax =  tuple(request.POST.getlist("tax[]"))
    pdebitid = purchasedebit.objects.get(pdebitid=pdebt.pdebitid, company=cmp)
   

 
    print('product==',product)
    print('qty==',qty)
    print('discount==',discount)
    print('total==',total) 
    if len(product)==len(qty)==len(discount)==len(total)==len(tax):
        mapped=zip(product,qty,discount,total,tax)
        mapped=list(mapped)
        for ele in mapped:
          itm = ItemModel.objects.get(id=ele[0])
          
          purchasedebit1.objects.create(product =itm,qty=ele[1],discount=ele[2],total=ele[3],tax=ele[4],pdebit=pdebitid,company=cmp)

    purchasedebit.objects.filter(company=cmp).update(tot_debt_no=F('tot_debt_no') + 1)
          
    pdebt.tot_debt_no = pdebt.pdebitid
    pdebt.save()

    DebitnoteTransactionHistory.objects.create(debitnote=pdebt,staff=staff,company=cmp,action='Created')

    if 'Next' in request.POST:
      return redirect('add_debitnote')
    
    if "Save" in request.POST:
      return redirect('view_purchasedebit')
    
  else:
    return render(request,'adddebitnotes.html',context)
    
def view_purchasedebit(request):
  if 'staff_id' in request.session:
    if request.session.has_key('staff_id'):
      staff_id = request.session['staff_id']
           
    else:
      return redirect('/')
  staff =  staff_details.objects.get(id=staff_id)
  cmp = company.objects.get(id=staff.company.id)
  print("hello")
  print(staff)
  allmodules = modules_list.objects.filter(company=staff.company, status='New').first()
  pdebt = purchasedebit.objects.filter(company=cmp)

  if not pdebt:
    context = {'staff':staff, 'allmodules':allmodules}
    return render(request,'emptydebit.html',context)
  
  context = {'staff':staff,'allmodules':allmodules,'pdebt':pdebt}
  return render(request,'purchase_return_dr.html',context)
    

def view_purchasedebit(request):
  sid = request.session.get('staff_id')
  staff =  staff_details.objects.get(id=sid)
  cmp = company.objects.get(id=staff.company.id)
  pdebt = purchasedebit.objects.filter(company=cmp)
  
  if not pdebt:
    context = {'staff':staff}
    return render(request,'emptydebit.html',context)
  
  context = {'staff':staff,'pdebt':pdebt}
  return render(request,'purchase_return_dr.html',context)
  
def view_parties(request):
  staff_id = request.session['staff_id']
  staff =  staff_details.objects.get(id=staff_id)
  
 
  Party=party.objects.filter(company=staff.company.id)
  allmodules= modules_list.objects.get(company=staff.company,status='New')
  return render(request, 'view_parties.html',{'staff':staff,'allmodules':allmodules,'Party':Party})

def view_party(request,id):
  staff_id = request.session['staff_id']
  staff =  staff_details.objects.get(id=staff_id)
  getparty=party.objects.get(id=id)
  Party=party.objects.filter(company=staff.company.id)
  allmodules= modules_list.objects.get(company=staff.company,status='New')
  return render(request, 'view_party.html',{'staff':staff,'allmodules':allmodules,'Party':Party,'getparty':getparty})

def saveitem(request):
    sid = request.session.get('staff_id')
    staff =  staff_details.objects.get(id=sid)
    company_obj = staff.company

    name = request.POST['name']
    unit = request.POST['unit']
    hsn = request.POST['hsn']
    taxref = request.POST['taxref']
    sell_price = request.POST['sell_price']
    cost_price = request.POST['cost_price']
    intra_st = request.POST['intra_st']
    inter_st = request.POST['inter_st']

    if taxref != 'Taxable':
        intra_st = 'GST0[0%]'
        inter_st = 'IGST0[0%]'

    itmdate = request.POST.get('itmdate')
    stock = request.POST.get('stock')
    itmprice = request.POST.get('itmprice')
    minstock = request.POST.get('minstock')

    if not hsn:
        hsn = None

    # Check if HSN exists for the given company
    if ItemModel.objects.filter(company=company_obj, item_hsn=hsn).exists():
        return JsonResponse({'success': False, 'message': 'HSN already exists for this company'})

    itm = ItemModel(
        item_name=name, item_hsn=hsn, item_unit=unit, item_taxable=taxref, item_gst=intra_st, item_igst=inter_st,
        item_sale_price=sell_price, item_purchase_price=cost_price, item_current_stock=stock, item_at_price=itmprice,
        item_date=itmdate, company=company_obj, user=company_obj.user
    )
    itm.save()

    return JsonResponse({'success': True})

def itemdetails(request):
    itmid = request.GET['id']
    itm = ItemModel.objects.get(id=itmid)
    hsn = itm.item_hsn
    gst = itm.item_gst
    igst = itm.item_igst
    price = itm.item_purchase_price
    qty = itm.item_current_stock
    return JsonResponse({'hsn': hsn, 'gst': gst, 'igst': igst, 'price': price, 'qty': qty})



# def savecustomer1(request):
#     sid = request.session.get('staff_id')
#     staff = staff_details.objects.get(id=sid)
#     cmp = company.objects.get(id=staff.company.id)

#     party_name = request.POST['name']
#     email = request.POST['email']
    
#     phone_number = request.POST.get('mobile', None)
#     state = request.POST['splystate']
#     address = request.POST['baddress']
#     gst_type = request.POST['gsttype']
#     gst_no = request.POST['gstin']
#     openingbalance = request.POST.get('openbalance')
#     payment = request.POST.get('paytype')
#     End_date = request.POST.get('enddate', None)
    
    
#     part = Parties(
#         party_name=party_name,
#         gstin=gst_no,
#         phone_number=phone_number,
#         gst_type=gst_type,
#         state=state,
#         billing_address=address,
#         email=email,
#         opening_balance=openingbalance,
#         to_pay=(payment == 'to_pay'),  # Set to True if payment is 'to_pay'
#         to_recieve=(payment == 'to_receive'),        
#         date=End_date,
#         company=cmp,
#         staff=staff,

        
#     )
#     part.save()
#     history_action =  "Created"
#     history = History(
#         staff=staff,
#         company=cmp,
#         party=part,
#         action=history_action,
#     )
#     history.save()
    
  
    
#     partyobj = Parties.objects.filter(company=cmp).values('id', 'party_name')

#     party_list = [{'id': customer['id'], 'name': customer['party_name']} for customer in partyobj]

#     return JsonResponse({'party_list':party_list, 'success': True, 'message': 'Customer saved successfully.'})
  
def savecustomer1(request):
    if request.method == 'POST':
        sid = request.session.get('staff_id')
        staff = staff_details.objects.get(id=sid)
        cmp = company.objects.get(id=staff.company.id)

        party_name = request.POST['name']
        email = request.POST['email']
        contact = request.POST['contact']
        state = request.POST['splystate']
        billing_address = request.POST['baddress']
        gst_type = request.POST['gsttype']
        gstin = request.POST['gstin']
        openingbalance = request.POST.get('openbalance')
        End_date = request.POST.get('enddate', None)
       

        

        # Check if contact already exists
        if Parties.objects.filter(phone_number=contact,company=cmp).exists():
            print("phone;")
            messages.info(request, 'Sorry, Contact Number already exists')
            return redirect('add_debitnote')
        
        elif Parties.objects.filter(email=email, company=cmp).exists():
            print("email;")
            messages.info(request, 'Sorry, Email already exists')
            return redirect('add_debitnote')
        # elif Parties.objects.filter(gstin=gstin, company=cmp).exists():
        #     print("gst;")
        #     messages.info(request, 'Sorry, GST number already exists')
        #     return redirect('add_debitnote')
        else:
            print("saved;")
            part = Parties(party_name=party_name, gstin=gstin, phone_number=contact, gst_type=gst_type, state=state, billing_address=billing_address,
                         email=email, opening_balance=openingbalance,
                         date=End_date,company=cmp,staff=staff)

            part.save()
            print(part,'party')
            return JsonResponse({'success': True, 'id':part.id})
    else:
        return JsonResponse({'success': False, 'message': 'Invalid request method'})

def cust_dropdown1(request):
  sid = request.session.get('staff_id')
  staff =  staff_details.objects.get(id=sid)
  cmp = company.objects.get(id=staff.company.id)
  part = Parties.objects.filter(company=cmp) 

  id_list = []
  party_list = []
  for p in part:
    id_list.append(p.id)
    party_list.append(p.party_name)

  return JsonResponse({'id_list':id_list, 'party_list':party_list })


# def cust_dropdown1(request):
#   sid = request.session.get('staff_id')
#   staff =  staff_details.objects.get(id=sid)
#   cmp = company.objects.get(id=staff.company.id)
#   part = Parties.objects.filter(company=cmp)

#   id_list = []
#   party_list = []
#   for c in part:
#     id_list.append(c.id)
#     party_list.append(c.party_name)

#   return JsonResponse({'id_list':id_list, 'party_list':party_list })


# def saveitem1(request):
#     sid = request.session.get('staff_id')
#     staff =  staff_details.objects.get(id=sid)
#     company_obj = staff.company

#     name = request.POST['name']
#     unit = request.POST['unit']
#     hsn = request.POST['hsn']
#     taxref = request.POST['taxref']
#     sell_price = request.POST['sell_price']
#     cost_price = request.POST['cost_price']
#     intra_st = request.POST['intra_st']
#     inter_st = request.POST['inter_st']

#     if taxref != 'Taxable':
#         intra_st = 'GST0[0%]'
#         inter_st = 'IGST0[0%]'

#     itmdate = request.POST.get('itmdate')
#     stock = request.POST.get('stock')
#     itmprice = request.POST.get('itmprice')
#     minstock = request.POST.get('minstock')

#     if not hsn:
#         hsn = None

#     # Check if HSN exists for the given company
#     if ItemModel.objects.filter(company=company_obj, item_hsn=hsn).exists():
#         return JsonResponse({'success': False, 'message': 'HSN already exists for this company'})

#     itm = ItemModel(
#         item_name=name, item_hsn=hsn, item_unit=unit, item_taxable=taxref, item_gst=intra_st, item_igst=inter_st,
#         item_sale_price=sell_price, item_purchase_price=cost_price, item_current_stock=stock, item_at_price=itmprice,
#         item_date=itmdate, company=company_obj, user=company_obj.user
#     )
#     itm.save()

#     return JsonResponse({'success': True})
  
def saveitem1(request):
  if request.method == 'POST':
    sid = request.session.get('staff_id')
    staff =  staff_details.objects.get(id=sid)
    cmp = company.objects.get(id=staff.company.id)

    name = request.POST['name']
    unit = request.POST['unit']
    hsn = request.POST['hsn']
    taxref = request.POST['taxref']
    sell_price = request.POST['sell_price']
    cost_price = request.POST['cost_price']
    intra_st = request.POST['intra_st']
    inter_st = request.POST['inter_st']

    if taxref != 'Taxable':
        intra_st = 'GST0[0%]'
        inter_st = 'IGST0[0%]'

    itmdate = request.POST.get('itmdate')
    stock = request.POST.get('stock')
    itmprice = request.POST.get('itmprice')
    minstock = request.POST.get('minstock')

    # Check if the HSN already exists
    if ItemModel.objects.filter(item_hsn=hsn,company=cmp).exists():
       messages.info(request, 'Sorry, HSN Number already exists')
       return redirect('add_debitnote')
    if ItemModel.objects.filter(item_name=name,company=cmp).exists():
       messages.info(request, 'Sorry, ITEM Name already exists')
       return redirect('add_debitnote')
    else:
        itm = ItemModel(item_name=name,item_hsn=hsn,item_unit=unit,item_taxable=taxref, item_gst=intra_st,item_igst=inter_st, item_sale_price=sell_price, 
                    item_purchase_price=cost_price,item_opening_stock=stock,item_current_stock=stock,item_at_price=itmprice,item_date=itmdate,
                    item_min_stock_maintain=minstock,company=cmp,user=cmp.user)
        itm.save() 
        return JsonResponse({'success': True})
    

        
def item_dropdowns(request):
  sid = request.session.get('staff_id')
  staff =  staff_details.objects.get(id=sid)
  cmp = company.objects.get(id=staff.company.id)
  product = ItemModel.objects.filter(company=cmp,user=cmp.user)

  id_list = []
  product_list = []
  items = {}
  for p in product:
    id_list.append(p.id)
    product_list.append(p.item_name)
    items[p.id] = [p.id, p.item_name]
  return JsonResponse(items)

# def custdata1(request):
#     cid = request.GET.get('customer_id')
#     part = Parties.objects.get(id=cid)
#     phno = part.phone_number
#     address = part.billing_address
#     pay = 'To Pay' if part.to_pay else 'To Receive'
#     bal = part.opening_balance
#     return JsonResponse({'phno': phno, 'address': address, 'pay': pay, 'bal': bal})

def custdata1(request):
  cid = request.POST['id']
  part = Parties.objects.get(id=cid)
  # email = part.email
  phno = part.phone_number
  address = part.billing_address
  # pay = part.payment
  
  bal = part.opening_balance
  
  
  # bill=[]
  # p=PurchaseBill.objects.filter(party=part).values('billno')
  # for i in p:
  #    bill.append(i)


  try:
    b=[bill for bill in PurchaseBill.objects.filter(party=part).values('billno')]
  except:
    b=None
  print(b,'billno')
  return JsonResponse({ 'phno':phno, 'address':address, 'bal':bal,'billno':b})

def purchasebilldata(request):
    try:
        party_name = request.POST['id']
        party_instance = Parties.objects.get(id=party_name)

        # Initialize lists to store multiple bill numbers and dates
        bill_numbers = []
        bill_dates = []

        try:
            # Retrieve all PurchaseBill instances for the party
            bill_instances = PurchaseBill.objects.filter(party=party_instance)

            # Loop through each PurchaseBill instance and collect bill numbers and dates
            for bill_instance in bill_instances:
                bill_numbers.append(bill_instance.billno)
                bill_dates.append(bill_instance.billdate)

        except PurchaseBill.DoesNotExist:
            pass

        # Return a JSON response with the list of bill numbers and dates
        if not bill_numbers and not bill_dates:
            return JsonResponse({'bill_numbers': ['nobill'], 'bill_dates': ['nodate']})

        return JsonResponse({'bill_numbers': bill_numbers, 'bill_dates': bill_dates})

    except KeyError:
        return JsonResponse({'error': 'The key "id" is missing in the POST request.'})

    except party.DoesNotExist:
        return JsonResponse({'error': 'Party not found.'})
    
def get_bill_date(request):
    selected_bill_no = request.POST.get('bill_no', None)



    try:
        
        # Get the latest PurchaseBill with the specified bill_number
    
        purchase_bill = PurchaseBill.objects.filter(billno=selected_bill_no).latest('billdate')
        part=purchase_bill.party
        phno = part.phone_number
        address = part.billing_address
        customername=part.party_name
  # pay = part.payment
        bal = part.opening_balance
        bill_date = purchase_bill.billdate.strftime('%Y-%m-%d')
 
    except PurchaseBill.DoesNotExist:
        return JsonResponse({'error': 'Bill number not found'}, status=400)
    except PurchaseBill.MultipleObjectsReturned:
        # Handle the case where multiple PurchaseBills are found for the same bill_number
        return JsonResponse({'error': 'Multiple PurchaseBills found for the same bill number'}, status=400)

    return JsonResponse({'bill_date': bill_date,'phno':phno,'address':address,'bal':bal,'customername':customername})


def bankdata1(request):
  bid = request.POST['id']
  bank = BankModel.objects.get(id=bid) 
  bank_no = bank.account_num
  return JsonResponse({'bank_no':bank_no})
  
  
def parties_default(request):
  sid = request.session.get('staff_id')
  staff =  staff_details.objects.get(id=sid)
  cmp = company.objects.get(id=staff.company.id)

  return render(request, 'parties_default.html',{'staff':staff, 'cmp':cmp})
   
   
# Harikrishnan ---------------------------------------

def parties_add_page(request):
    todaydate = date.today().isoformat()
    return render(request,'parties_add_page.html',{'todaydate':todaydate})



def parties_table(request):
    sid = request.session.get('staff_id')
    staff = staff_details.objects.get(id=sid)
    cmp = company.objects.get(id=staff.company.id)
    comp = cmp.id
    count = Parties.objects.filter(company_id=cmp.id).count()
    if count > 0:
        transactions = Parties.objects.all().order_by('date')
        for index, transaction in enumerate(transactions, start=1):
          transaction.index = index
        parties = Parties.objects.filter(company_id=comp).values('party_name','phone_number').annotate(total_amount=Sum('opening_balance')).distinct().order_by('party_name')
        
        return render(request,'parties_table.html',{'parties':parties,'transactions':transactions})
    else:
        return render(request,'parties_default.html')
    

    
def party_save(request):
    
    sid = request.session.get('staff_id')
    staff = staff_details.objects.get(id=sid)
    cmp = company.objects.get(id=staff.company.id)
    user = cmp.id
    
    if request.method == 'POST':
        partyname = request.POST['partyname'].capitalize()
        mobilenumber = request.POST['mobilenumber']
        gstin = request.POST['gstin']
        gstintype = request.POST['gstintype']
        state = request.POST['state']
        email = request.POST['email']
        Date = request.POST['date']
        address = request.POST['address']
        balance = request.POST['balance']
        buttonn = request.POST['buttonn']

        if Parties.objects.filter(party_name = partyname, phone_number = mobilenumber,company_id=user).exists():
          messages.error(request, 'User already exists!')
          todaydate = date.today().isoformat()
          return render(request,'parties_add_page.html',{'todaydate':todaydate})
        
        elif Parties.objects.filter(phone_number = mobilenumber,company_id=user).exists():
            messages.error(request, 'Phone number already exists!')
            todaydate = date.today().isoformat()
            return render(request,'parties_add_page.html',{'todaydate':todaydate})
        
        elif Parties.objects.filter(party_name = partyname,company_id=user).exists():
            messages.error(request, 'Identical name exists! Please use initials.')
            todaydate = date.today().isoformat()
            return render(request,'parties_add_page.html',{'todaydate':todaydate})
           
        else:
          if balance == '' or balance == '0' :
                 
              party = Parties(party_name = partyname,phone_number = mobilenumber, gstin = gstin,
                              gst_type = gstintype, billing_address = address, state = state,
                              email = email, date = Date,company_id=user,staff_id=staff.id)
              party.save()

              history = History(company_id=user,party_id=party.id,staff_id=staff.id,action='CREATED')
              history.save()
              
          else:
              if request.POST['pay_recieve'] != '':
                  pay_recieve = request.POST['pay_recieve']

                  if pay_recieve == 'receive':
                      party = Parties(party_name = partyname,phone_number = mobilenumber, gstin = gstin,
                                      gst_type = gstintype, billing_address = address, state = state,
                                      email = email, date = Date,opening_balance = balance,to_recieve = True,
                                      company_id=user,staff_id=staff.id)
                      party.save()
                      history = History(company_id=user,party_id=party.id,staff_id=staff.id,action='CREATED')
                      history.save()

                  elif pay_recieve == 'pay':
                      neg_balance = -int(balance)
                      party = Parties(party_name = partyname,phone_number = mobilenumber, gstin = gstin,
                                      gst_type = gstintype, billing_address = address, state = state,
                                      email = email, date = Date,opening_balance = neg_balance,to_pay = True,
                                      company_id=user,staff_id=staff.id)
                      party.save()
                      history = History(company_id=user,party_id=party.id,staff_id=staff.id,action='CREATED')
                      history.save()
                  else:
                      party = Parties(party_name = partyname,phone_number = mobilenumber, gstin = gstin,
                                      gst_type = gstintype, billing_address = address, state = state,
                                      email = email, date = Date,staff_id=staff.id,company_id=user)
                      party.save()
                      history = History(company_id=user,party_id=party.id,staff_id=staff.id,action='CREATED')
                      history.save()

          if buttonn == 'new':
              return redirect('parties_add_page')
          elif buttonn == 'old':
              return redirect('parties_table')


def party_delete(request,pk):
    sid = request.session.get('staff_id')
    staff = staff_details.objects.get(id=sid)
    cmp = company.objects.get(id=staff.company.id)
    user = cmp.id
    party = Parties.objects.filter(phone_number=pk,company_id=user).first()
    deleteparty = Parties.objects.filter(phone_number=party.phone_number,party_name=party.party_name)
    deleteparty.delete()
    return redirect('parties_table')


def party_edit(request,pk,id):
    sid = request.session.get('staff_id')
    staff = staff_details.objects.get(id=sid)
    cmp = company.objects.get(id=staff.company.id)
    user = cmp.id
    party = get_object_or_404(Parties, phone_number=pk,party_name=id,company_id=user)
    return render(request,'parties_edit.html',{ 'partyy':party })



def party_update(request,pk):

    sid = request.session.get('staff_id')
    staff = staff_details.objects.get(id=sid)
    cmp = company.objects.get(id=staff.company.id)
    party = Parties.objects.get(id=pk)
    user = cmp.id
    if request.method == 'POST':
        
        partyname = request.POST['partyname'].capitalize()
        mobile = request.POST['mobilenumber']

        if Parties.objects.filter(party_name = partyname, phone_number = mobile,company_id=user).exclude(id=pk).exists():
          messages.error(request, 'User already exists!')
          todaydate = date.today().isoformat()
          partyy = get_object_or_404(Parties, phone_number=party.phone_number,party_name=party.party_name,company_id=user)
          return render(request,'parties_edit.html',{'todaydate':todaydate,'partyy':partyy})
        
        elif Parties.objects.filter(phone_number = mobile,company_id=user).exclude(id=pk).exists():
            messages.error(request, 'Phone number already exists!')
            todaydate = date.today().isoformat()
            partyy = get_object_or_404(Parties, phone_number=party.phone_number,party_name=party.party_name,company_id=user)
            return render(request,'parties_edit.html',{'todaydate':todaydate,'partyy':partyy})
        
        elif Parties.objects.filter(party_name = partyname,company_id=user).exclude(id=pk).exists():
            messages.error(request, 'Identical name exists! Please use initials.')
            todaydate = date.today().isoformat()
            partyy = get_object_or_404(Parties, phone_number=party.phone_number,party_name=party.party_name,company_id=user)
            return render(request,'parties_edit.html',{'todaydate':todaydate,'partyy':partyy})
           
        else:

          party.party_name = request.POST['partyname'].capitalize()
          party.phone_number = request.POST['mobilenumber']
          party.gstin = request.POST['gstin']
          party.gst_type = request.POST['gstintype']
          party.state = request.POST['state']
          party.email = request.POST['email']
          party.date = request.POST['date']
          party.billing_address = request.POST['address']

          if float(request.POST['balance']) > 0:
              party.opening_balance = request.POST['balance']
              if request.POST['pay_recieve'] == 'pay':
                party.to_pay = True
                party.to_recieve = False
                neg_balance = request.POST['balance']
                party.opening_balance = -float(neg_balance)
                
              else:
                party.to_pay = False
                party.to_recieve = True
                party.opening_balance = request.POST['balance']
          else :
            party.opening_balance = 0
          
          party.save()
          history = History(company_id=user,party_id=pk,staff_id=sid,action='UPDATED')
          history.save()
          
          return redirect('parties_table')
    
def parties_history(request,pk,id):
   sid = request.session.get('staff_id')
   staff = staff_details.objects.get(id=sid)
   cmp = company.objects.get(id=staff.company.id)
   user = cmp.id
   party= Parties.objects.get(phone_number=pk,party_name=id,company_id=user)
   history = History.objects.filter(party_id=party.id).order_by('date')
   return render(request,'parties_history.html',{'history':history,'partyName':party})

def party_details(request,pk,id):
   sid = request.session.get('staff_id')
   staff = staff_details.objects.get(id=sid)
   cmp = company.objects.get(id=staff.company.id)
   user = cmp.id
   party = Parties.objects.filter(phone_number=pk,party_name=id,company_id=user).order_by('date')
   details =  Parties.objects.get(phone_number=pk,party_name=id,company_id=user) 
   transactions = Parties.objects.all()
   for index, part in enumerate(party, start=1):
      part.index = index
   return render(request,'parties_details.html',{'party':party, 'details':details})

# end ----------------


def import_debitnote(request):
  if request.method == 'POST' and request.FILES['billfile']  and request.FILES['prdfile']:
    sid = request.session.get('staff_id')
    staff =  staff_details.objects.get(id=sid)
    cmp = company.objects.get(id=staff.company.id)
    totval = int(purchasedebit.objects.filter(company=cmp).last().tot_debt_no)

    excel_bill = request.FILES['billfile']
    excel_b = load_workbook(excel_bill)
    eb = excel_b['Sheet1']
    excel_prd = request.FILES['prdfile']
    excel_p = load_workbook(excel_prd)
    ep = excel_p['Sheet1']

    for row_number1 in range(2, eb.max_row + 1):
      debitsheet = [eb.cell(row=row_number1, column=col_num).value for col_num in range(1, eb.max_column + 1)]
      part = party.objects.get(party_name=debitsheet[0],email=debitsheet[1],company=cmp)
      purchasedebit.objects.create(party=part,pdebitid = totval,
                                  debitdate=debitsheet[2],
                                  supply =debitsheet[3],
                                  tot_debt_no = totval,
                                  company=cmp,staff=staff)
      
      pdebt = purchasedebit.objects.last()
      if debitsheet[4] == 'Cheque':
        pdebt.payment_type = 'Cheque'
        pdebt.cheque_no = debitsheet[5]
      elif debitsheet[4] == 'UPI':
        pdebt.upi_no = debitsheet[5]
      else:
        if debitsheet[4] != 'Cash':
          bank = BankModel.objects.get(bank_name=debitsheet[4],company=cmp)
          pdebt.payment_type = bank
        else:
          pdebt.payment_type = 'Cash'
      pdebt.save()

      purchasedebit.objects.filter(company=cmp).update(tot_debt_no=totval )
      totval += 1
      subtotal = 0
      taxamount=0
      for row_number2 in range(2, ep.max_row + 1):
        prdsheet = [ep.cell(row=row_number2, column=col_num).value for col_num in range(1, ep.max_column + 1)]
        if prdsheet[0] == row_number1:
          itm = ItemModel.objects.get(item_name=prdsheet[1],item_hsn=prdsheet[2],company=cmp)
          total=int(prdsheet[3])*int(itm.item_purchase_price) - int(prdsheet[4])
          
          purchasedebit1.objects.create(pdebit=pdebt,
                                company=cmp,
                                product=itm,
                                qty=prdsheet[3],
                                discount=prdsheet[4],
                                total=total)

       
          if debitsheet[3] =='State':
            taxval = itm.item_gst
            taxval=taxval.split('[')
            tax=int(taxval[0][3:])
          else:
            taxval = itm.item_igst
            taxval=taxval.split('[')
            tax=int(taxval[0][4:])

          subtotal += total
          tamount = total *(tax / 100)
          taxamount += tamount
                
          if debitsheet[3]=='State':
            gst = round((taxamount/2),2)
            pdebt.sgst=gst
            pdebt.cgst=gst
            pdebt.igst=0

          else:
            gst=round(taxamount,2)
            pdebt.igst=gst
            pdebt.cgst=0
            pdebt.sgst=0

      gtotal = subtotal + taxamount + float(debitsheet[6])
      balance = gtotal- float(debitsheet[7])
      gtotal = round(gtotal,2)
      balance = round(balance,2)

      pdebt.subtotal=round(subtotal,2)
      pdebt.taxamount=round(taxamount,2)
      pdebt.adjustment=round(debitsheet[6],2)
      pdebt.grandtotal=gtotal
      pdebt.paid_amount=round(debitsheet[7],2)
      pdebt.balance_amount=balance
      pdebt.save()

      DebitnoteTransactionHistory.objects.create(debitnote=pdebt,staff=pdebt.staff,company=pdebt.company,action='Created')
      return JsonResponse({'message': 'File uploaded successfully!'})
  else:
    return JsonResponse({'message': 'File upload Failed!'})
    
    
def import_purchase_bill(request):
  if request.method == 'POST' and 'billfile' in request.FILES and 'prdfile' in request.FILES:
    sid = request.session.get('staff_id')
    staff =  staff_details.objects.get(id=sid)
    cmp = company.objects.get(id=staff.company.id)
    totval = int(PurchaseBill.objects.filter(company=cmp).last().tot_bill_no) + 1

    excel_bill = request.FILES['billfile']
    excel_b = load_workbook(excel_bill)
    eb = excel_b['Sheet1']
    excel_prd = request.FILES['prdfile']
    excel_p = load_workbook(excel_prd)
    ep = excel_p['Sheet1']

    for row_number1 in range(2, eb.max_row + 1):
      billsheet = [eb.cell(row=row_number1, column=col_num).value for col_num in range(1, eb.max_column + 1)]
      part = Parties.objects.get(party_name=billsheet[0],email=billsheet[1],company=cmp)
      PurchaseBill.objects.create(party=part,billno=totval,
                                  billdate=billsheet[2],
                                  supplyplace =billsheet[3],
                                  tot_bill_no = totval,
                                  company=cmp,staff=staff)
      
      pbill = PurchaseBill.objects.last()
      if billsheet[4] == 'Cheque':
        pbill.pay_method = 'Cheque'
        pbill.cheque_no = billsheet[5]
      elif billsheet[4] == 'UPI':
        pbill.pay_method = 'UPI'
        pbill.upi_no = billsheet[5]
      else:
       
        pbill.save()

      PurchaseBill.objects.filter(company=cmp).update(tot_bill_no=totval)
      totval += 1
      subtotal = 0
      taxamount=0
      for row_number2 in range(2, ep.max_row + 1):
        prdsheet = [ep.cell(row=row_number2, column=col_num).value for col_num in range(1, ep.max_column + 1)]
        if prdsheet[0] == row_number1:
          itm = ItemModel.objects.get(item_name=prdsheet[1],item_hsn=int(prdsheet[2]),company=cmp)
          total=int(prdsheet[3])*int(itm.item_purchase_price) - int(prdsheet[4])
          PurchaseBillItem.objects.create(purchasebill=pbill,
                                company=cmp,
                                product=itm,
                                qty=prdsheet[3],
                                discount=prdsheet[4],
                                total=total)

          if billsheet[3] =='State':
            taxval = itm.item_gst
            taxval=taxval.split('[')
            tax=int(taxval[0][3:])
          else:
            taxval = itm.item_igst
            taxval=taxval.split('[')
            tax=int(taxval[0][4:])

          subtotal += total
          tamount = total *(tax / 100)
          taxamount += tamount
                
          if billsheet[3]=='State':
            gst = round((taxamount/2),2)
            pbill.sgst=gst
            pbill.cgst=gst
            pbill.igst=0

          else:
            gst=round(taxamount,2)
            pbill.igst=gst
            pbill.cgst=0
            pbill.sgst=0

      gtotal = subtotal + taxamount + float(billsheet[6])
      balance = gtotal- float(billsheet[7])
      gtotal = round(gtotal,2)
      balance = round(balance,2)

      pbill.subtotal=round(subtotal,2)
      pbill.taxamount=round(taxamount,2)
      pbill.adjust=round(billsheet[6],2)
      pbill.grandtotal=gtotal
      pbill.advance=round(billsheet[7],2)
      pbill.balance=balance
      pbill.save()

      PurchaseBillTransactionHistory.objects.create(purchasebill=pbill,staff=pbill.staff,company=pbill.company,action='Created')
      return JsonResponse({'message': 'File uploaded successfully!'})
  else:
    return JsonResponse({'message': 'File upload Failed!'})
    
    
def view_purchasebill(request):
  sid = request.session.get('staff_id')
  staff =  staff_details.objects.get(id=sid)
  cmp = company.objects.get(id=staff.company.id)
  pbill = PurchaseBill.objects.filter(company=cmp)
  
  if not pbill:
    context = {'staff':staff}
    return render(request,'purchasebillempty.html',context)
  
  context = {'staff':staff,'pbill':pbill}
  return render(request,'purchasebilllist.html',context)


def add_purchasebill(request):
  toda = date.today()
  tod = toda.strftime("%Y-%m-%d")

  sid = request.session.get('staff_id')
  staff =  staff_details.objects.get(id=sid)
  cmp = company.objects.get(id=staff.company.id)
  cust = Parties.objects.filter(company=cmp)
  pbill = PurchaseBill.objects.filter(company=cmp)

  
  
  
  last_bill = PurchaseBill.objects.filter(company=cmp).last()

  if last_bill:
    bill_no = last_bill.tot_bill_no + 1 
  else:
    bill_no = 1

  item = ItemModel.objects.filter(company=cmp,user=cmp.user)

  context = {'staff':staff, 'cust':cust, 'cmp':cmp,'bill_no':bill_no, 'tod':tod, 'item':item, }
  return render(request,'purchasebilladd.html',context)


def show_purchasebill(request,id):
    purchase_bill = get_object_or_404(PurchaseBill, id=id)

 
    pitm = PurchaseBillItem.objects.filter(purchasebill=purchase_bill)

   
    dis = sum(int(itm.discount) for itm in pitm)
    itm_len = len(pitm)

    context = {'purchase_bill': purchase_bill, 'pitm': pitm, 'itm_len': itm_len, 'dis': dis}
    return render(request, 'purchasebilldetails.html', context)

def create_purchasebill(request):
    if request.method == 'POST':
        sid = request.session.get('staff_id')
        staff = staff_details.objects.get(id=sid)
        cmp = company.objects.get(id=staff.company.id)
        part = Parties.objects.get(id=request.POST.get('customername'))
        pbill = PurchaseBill(party=part,
                              billno=request.POST.get('bill_no'),
                              billdate=request.POST.get('billdate'),
                              supplyplace=request.POST.get('placosupply'),
                              advance=request.POST.get("advance"),
                              balance=request.POST.get("balance"),
                              subtotal=float(request.POST.get('subtotal')),
                              igst=request.POST.get('igst'),
                              cgst=request.POST.get('cgst'),
                              sgst=request.POST.get('sgst'),
                              adjust=request.POST.get("adj"),
                              taxamount=request.POST.get("taxamount"),
                              grandtotal=request.POST.get('grandtotal'),
                              company=cmp, staff=staff)
        pbill.save()

        product = tuple(request.POST.getlist("product[]"))
        qty = tuple(request.POST.getlist("qty[]"))
        discount = tuple(request.POST.getlist("discount[]"))
        total = tuple(request.POST.getlist("total[]"))

        billno_instance = PurchaseBill.objects.filter(billno=pbill.billno, company=cmp).first()

        if len(product) == len(qty) == len(discount) == len(total) and billno_instance:
            mapped = zip(product, qty, discount, total)
            mapped = list(mapped)
            for ele in mapped:
                itm = ItemModel.objects.get(id=ele[0])
                PurchaseBillItem.objects.create(product=itm, qty=ele[1], discount=ele[2], total=ele[3],
                                                purchasebill=billno_instance, company=cmp)
       

        PurchaseBill.objects.filter(company=cmp).update(tot_bill_no=F('tot_bill_no') + 1)

        pbill.tot_bill_no = pbill.billno
        pbill.save()

        PurchaseBillTransactionHistory.objects.create(purchasebill=pbill, company=cmp, staff=staff, action='Created')

        if 'Next' in request.POST:
            return redirect('add_purchasebill')

        if "Save" in request.POST:
            return redirect('view_purchasebill')

    else:
        return render(request, 'purchasebilladd.html')

def savecustomer(request):
    sid = request.session.get('staff_id')
    staff = staff_details.objects.get(id=sid)
    cmp = company.objects.get(id=staff.company.id)

    party_name = request.POST['name']
    email = request.POST['email']
    
    phone_number = request.POST.get('mobile', None)
    state = request.POST['splystate']
    address = request.POST['baddress']
    gst_type = request.POST['gsttype']
    gst_no = request.POST['gstin']
    openingbalance = request.POST.get('openbalance')
    payment = request.POST.get('paytype')
    End_date = request.POST.get('enddate', None)
    
    
    part = Parties(
        party_name=party_name,
        gstin=gst_no,
        phone_number=phone_number,
        gst_type=gst_type,
        state=state,
        billing_address=address,
        email=email,
        opening_balance=openingbalance,
        to_pay=(payment == 'to_pay'),  # Set to True if payment is 'to_pay'
        to_recieve=(payment == 'to_receive'),        
        date=End_date,
        company=cmp,
        staff=staff,

        
    )
    part.save()
    history_action =  "Created"
    history = History(
        staff=staff,
        company=cmp,
        party=part,
        action=history_action,
    )
    history.save()
    
  
    
    partyobj = Parties.objects.filter(company=cmp).values('id', 'party_name')

    party_list = [{'id': customer['id'], 'name': customer['party_name']} for customer in partyobj]

    return JsonResponse({'party_list':party_list, 'success': True, 'message': 'Customer saved successfully.'})
    
    
def itemdetail(request):
    itmid = request.GET['id']
    itm = ItemModel.objects.get(id=itmid)
    hsn = itm.item_hsn
    gst = itm.item_gst
    igst = itm.item_igst
    price = itm.item_purchase_price
    qty = itm.item_current_stock
    return JsonResponse({'hsn': hsn, 'gst': gst, 'igst': igst, 'price': price, 'qty': qty})


def cust_dropdown(request):
  sid = request.session.get('staff_id')
  staff =  staff_details.objects.get(id=sid)
  cmp = company.objects.get(id=staff.company.id)
  part = Parties.objects.filter(company=cmp)

  id_list = []
  party_list = []
  for c in part:
    id_list.append(c.id)
    party_list.append(c.party_name)

  return JsonResponse({'id_list':id_list, 'party_list':party_list })

  
    
def save_item(request):
    if request.method == 'POST':
        item_name = request.POST.get('item_name')
        hsn = request.POST.get('hsn')
        qty = request.POST.get('qty')
        tax_ref = request.POST.get('taxref')
        intra_st = request.POST.get('intra_st')
        inter_st = request.POST.get('inter_st')
        sale_price = request.POST.get('saleprice')
        purchase_price = request.POST.get('purprice')

        item = ItemModel(
            item_name=item_name,
            item_hsn=hsn,
            item_current_stock=qty,
            item_taxable=tax_ref,
            item_gst=intra_st,
            item_igst=inter_st,
            item_sale_price=sale_price,
            item_purchase_price=purchase_price
        )
        item.save()

        
        return redirect("add_purchasebill")

    return render(request, 'purchasebilladd.html')



def custdata(request):
    cid = request.GET.get('customer_id')
    part = Parties.objects.get(id=cid)
    phno = part.phone_number
    address = part.billing_address
    pay = 'To Pay' if part.to_pay else 'To Receive'
    bal = part.opening_balance
    return JsonResponse({'phno': phno, 'address': address, 'pay': pay, 'bal': bal})


def history(request):
  return render(request,'purchasebillhistory.html')

def item_dropdown1(request):
  sid = request.session.get('staff_id')
  staff =  staff_details.objects.get(id=sid)
  cmp = company.objects.get(id=staff.company.id)
  product = ItemModel.objects.filter(company=cmp,user=cmp.user)

  id_list = []
  product_list = []
  items = {}
  for p in product:
    id_list.append(p.id)
    product_list.append(p.item_name)
    items[p.id] = [p.id, p.item_name]
  return JsonResponse(items)


def item_dropdown(request):
  sid = request.session.get('staff_id')
  staff =  staff_details.objects.get(id=sid)
  cmp = company.objects.get(id=staff.company.id)
  product = ItemModel.objects.filter(company=cmp,user=cmp.user)

  id_list = []
  product_list = []
  items = {}
  for p in product:
    id_list.append(p.id)
    product_list.append(p.item_name)
    items[p.id] = [p.id, p.item_name]
  return JsonResponse(items)

def edit_purchasebill(request,id):
  toda = date.today()
  tod = toda.strftime("%Y-%m-%d")
  
  sid = request.session.get('staff_id')
  staff =  staff_details.objects.get(id=sid)
  cmp = company.objects.get(id=staff.company.id)
  cust = Parties.objects.filter(company=cmp)
  item = ItemModel.objects.filter(company=cmp,user=cmp.user)
 
  pbill = PurchaseBill.objects.get(id=id,company=cmp)
  billprd = PurchaseBillItem.objects.filter(purchasebill=pbill,company=cmp)

 
  bdate = pbill.billdate.strftime("%Y-%m-%d")
  context = {'staff':staff, 'pbill':pbill, 'billprd':billprd,'tod':tod,
             'cust':cust, 'item':item,  'bdate':bdate,'phone_number': pbill.party.phone_number,
           'billing_address': pbill.party.billing_address}
  return render(request,'purchasebilledit.html',context)


def update_purchasebill(request,id):
  if request.method =='POST':
    sid = request.session.get('staff_id')
    staff = staff_details.objects.get(id=sid)
    cmp = company.objects.get(id=staff.company.id)  
    part = Parties.objects.get(id=request.POST.get('customername'))
    pbill = PurchaseBill.objects.get(id=id,company=cmp)
    pbill.party = part
    pbill.billdate = request.POST.get('billdate')
    pbill.supplyplace  = request.POST.get('placosupply')
    pbill.subtotal =float(request.POST.get('subtotal'))
    pbill.grandtotal = request.POST.get('grandtotal')
    pbill.igst = request.POST.get('igst')
    pbill.cgst = request.POST.get('cgst')
    pbill.sgst = request.POST.get('sgst')
    pbill.taxamount = request.POST.get("taxamount")
    pbill.adjust = request.POST.get("adj")
   
    pbill.advance = request.POST.get("advance")
    pbill.balance = request.POST.get("balance")

    pbill.save()

    product = tuple(request.POST.getlist("product[]"))
    qty = tuple(request.POST.getlist("qty[]"))
    total = tuple(request.POST.getlist("total[]"))
    discount = tuple(request.POST.getlist("discount[]"))

    PurchaseBillItem.objects.filter(purchasebill=pbill,company=cmp).delete()
    if len(total)==len(discount)==len(qty):
      mapped=zip(product,qty,discount,total)
      mapped=list(mapped)
      for ele in mapped:
        itm = ItemModel.objects.get(id=ele[0])
        PurchaseBillItem.objects.create(product =itm,qty=ele[1],discount=ele[2],total=ele[3],purchasebill=pbill,company=cmp)

    PurchaseBillTransactionHistory.objects.create(purchasebill=pbill,company=cmp,staff=staff,action='Updated')
    return redirect('view_purchasebill')

  return redirect('view_purchasebill')


def details_purchasebill(request,id):
  sid = request.session.get('staff_id')
  staff = staff_details.objects.get(id=sid)
  cmp = company.objects.get(id=staff.company.id) 
  
  pbill = PurchaseBill.objects.get(id=id,company=cmp)
  pitm = PurchaseBillItem.objects.filter(purchasebill=pbill,company=cmp)
  dis = 0
  for itm in pitm:
    dis += int(itm.discount)
  itm_len = len(pitm)

  context={'staff':staff,'pbill':pbill,'pitm':pitm,'itm_len':itm_len,'dis':dis}
  return render(request,'purchasebilldetails.html',context)


def history_purchasebill(request,id):
  sid = request.session.get('staff_id')
  staff = staff_details.objects.get(id=sid)
  cmp = company.objects.get(id=staff.company.id)   
  
  pbill = PurchaseBill.objects.get(id=id,company=cmp)
  hst= PurchaseBillTransactionHistory.objects.filter(purchasebill=pbill,company=cmp)

  context = {'staff':staff,'hst':hst,'pbill':pbill}
  return render(request,'purchasebillhistory.html',context)


def delete_purchasebill(request,id):
  sid = request.session.get('staff_id')
  staff = staff_details.objects.get(id=sid)
  cmp = company.objects.get(id=staff.company.id) 
  pbill = PurchaseBill.objects.get(id=id)
  PurchaseBillItem.objects.filter(purchasebill=pbill,company=cmp).delete()
  pbill.delete()
  return redirect('view_purchasebill')



  
def sharepdftomail(request,id):
  if request.user:
        try:
            if request.method == 'POST':
                emails_string = request.POST['email_ids']

                # Split the string by commas and remove any leading or trailing whitespace
                emails_list = [email.strip() for email in emails_string.split(',')]
                email_message = request.POST['email_message']
                print(emails_list)
                sid = request.session.get('staff_id')
                staff = staff_details.objects.get(id=sid)
                cmp = company.objects.get(id=staff.company.id) 
                
                pbill = PurchaseBill.objects.get(id=id,company=cmp)
                pitm = PurchaseBillItem.objects.filter(purchasebill=pbill,company=cmp)
                context = {'pbill':pbill, 'cmp':cmp,'pitm':pitm}
                template_path = 'purchase_mail.html'
                template = get_template(template_path)

                html  = template.render(context)
                result = BytesIO()
                pdf = pisa.pisaDocument(BytesIO(html.encode("ISO-8859-1")), result)
                pdf = result.getvalue()
                filename = f'DEBIT NOTE - {pbill.id}.pdf'
                subject = f"DEBIT NOTE - {pbill.id}"
                email = EmailMessage(subject, f"Hi,\nPlease find the attached DEBIT NOTE - File-{pbill.id}. \n{email_message}\n\n--\nRegards,\n{cmp.company_name}\n{cmp.address}\n{cmp.state} - {cmp.country}\n{cmp.contact}", from_email=settings.EMAIL_HOST_USER, to=emails_list)
                email.attach(filename, pdf, "application/pdf")
                email.send(fail_silently=False)

                msg = messages.success(request, 'Debit note file has been shared via email successfully..!')
                return redirect('details_purchasebill', id=id)

        except Exception as e:
            print(e)
            messages.error(request, f'{e}')
            return redirect('details_purchasebill', id=id)

 
def billhistory(request):
  pid = request.POST['id']
  sid = request.session.get('staff_id')
  staff = staff_details.objects.get(id=sid)
  cmp = company.objects.get(id=staff.company.id) 
  pbill = PurchaseBill.objects.get(billno=pid,company=cmp)
  hst = PurchaseBillTransactionHistory.objects.filter(purchasebill=pbill,company=cmp).last()
  name = hst.staff.first_name + ' ' + hst.staff.last_name 
  action = hst.action
  return JsonResponse({'name':name,'action':action,'pid':pid})
  
  
def check_gstin_exists(request):
    gstin = request.GET.get('gstin')
    sid = request.session.get('staff_id')
    staff = staff_details.objects.get(id=sid)
    cmp = company.objects.get(id=staff.company.id)
    gsttype = request.GET.get('gsttype')  
    print(gsttype)

    if gsttype != "Unregistered/Consumers":
        if Parties.objects.filter(company=cmp, gstin=gstin).exists():
            return JsonResponse({'exists': True})
        return JsonResponse({'exists': False})

def check_phone_number_exists(request):
    phone_number = request.GET.get('phone_number')
    sid = request.session.get('staff_id')
    staff =  staff_details.objects.get(id=sid)
    cmp = company.objects.get(id=staff.company.id)
    if Parties.objects.filter(company=cmp,phone_number=phone_number).exists():
        return JsonResponse({'exists': True})
    return JsonResponse({'exists': False})

# def check_contact_exists(request):
#     phone_number = request.GET.get('phone_number')  
#     if party.objects.filter(phone_number=phone_number).exists():
#         return JsonResponse({'exists': True})
#     return JsonResponse({'exists':False})

def check_contact_exists(request):
    phone_number = request.GET.get('phone_number')
    sid = request.session.get('staff_id')
    staff =  staff_details.objects.get(id=sid)
    cmp = company.objects.get(id=staff.company.id)
    if Parties.objects.filter(company=cmp,phone_number=phone_number).exists():
        return JsonResponse({'exists': True})
    return JsonResponse({'exists': False})

def check_name_exists(request):
    item_name= request.GET.get('item_name')
    
    if ItemModel.objects.filter(item_name=item_name).exists():
        return JsonResponse({'exists': True})
    return JsonResponse({'exists':False})
    
def check_email_exists(request):
    email= request.GET.get('email')
    if Parties.objects.filter(email=email).exists():
        return JsonResponse({'exists': True})
    return JsonResponse({'exists':False})
    
    

def check_hsn_exists(request):
    hsn= request.GET.get('hsn')
    
    if ItemModel.objects.filter(item_hsn=hsn).exists():
        return JsonResponse({'exists': True})
    return JsonResponse({'exists':False})
    
    
def customerdata(request):
  cid = request.POST['id']
  part = Parties.objects.get(id=cid)
  phno = part.phone_number
  address = part.billing_address
  pay =  'To Pay' if part.to_pay else 'To Receive'
  bal = part.opening_balance
  return JsonResponse({'phno':phno, 'address':address, 'pay':pay,'bal':bal})



def add_unit1(request):
    if request.method == 'POST':
        sid = request.session.get('staff_id')
        staff =  staff_details.objects.get(id=sid)
        cmp = company.objects.get(id=staff.company.id)
        new_unit_name = request.POST.get('newUnit')

        if new_unit_name:
           
            new_unit_name = new_unit_name.upper()

            
            if not ItemUnitModel.objects.filter(unit_name=new_unit_name).exists():
                
                new_unit = ItemUnitModel(user=cmp.user,company=cmp,unit_name=new_unit_name)
                new_unit.save()

                messages.success(request, 'Unit added successfully')
                return JsonResponse({'status': 'success', 'message': 'Unit added successfully'})
            else:
                messages.error(request, 'Unit already exists')
                return JsonResponse({'status': 'error', 'message': 'Unit already exists'})
        else:
            messages.error(request, 'Unit name cannot be empty')
            return JsonResponse({'status': 'error', 'message': 'Unit name cannot be empty'})

    return redirect('create_debitnotes') 

# def add_unit2(request):
#     if request.method == 'POST':
#         sid = request.session.get('staff_id')
#         staff =  staff_details.objects.get(id=sid)
#         cmp = company.objects.get(id=staff.company.id)
#         new_unit_name = request.POST.get('newUnit')

#         if new_unit_name:
           
#             new_unit_name = new_unit_name.upper()

            
#             if not ItemUnitModel.objects.filter(unit_name=new_unit_name).exists():
                
#                 new_unit = ItemUnitModel(user=cmp.user,company=cmp,unit_name=new_unit_name)
#                 new_unit.save()

#                 messages.success(request, 'Unit added successfully')
#                 return JsonResponse({'status': 'success', 'message': 'Unit added successfully'})
#             else:
#                 messages.error(request, 'Unit already exists')
#                 return JsonResponse({'status': 'error', 'message': 'Unit already exists'})
#         else:
#             messages.error(request, 'Unit name cannot be empty')
#             return JsonResponse({'status': 'error', 'message': 'Unit name cannot be empty'})

#     return redirect('update_debitnote')
  
  
def add_unit(request):
    if request.method == 'POST':
        new_unit_name = request.POST.get('newUnit')

        if new_unit_name:
           
            new_unit_name = new_unit_name.upper()

            
            if not ItemUnitModel.objects.filter(unit_name=new_unit_name).exists():
                
                new_unit = ItemUnitModel(unit_name=new_unit_name)
                new_unit.save()

                messages.success(request, 'Unit added successfully')
                return JsonResponse({'status': 'success', 'message': 'Unit added successfully'})
            else:
                messages.error(request, 'Unit already exists')
                return JsonResponse({'status': 'error', 'message': 'Unit already exists'})
        else:
            messages.error(request, 'Unit name cannot be empty')
            return JsonResponse({'status': 'error', 'message': 'Unit name cannot be empty'})

    return redirect('create_purchasebill')  
    

    
def sales_invoice(request):
    if 'staff_id' in request.session:
        if request.session.has_key('staff_id'):
            staff_id = request.session['staff_id']
        else:
            return redirect('/')
        
        staff = staff_details.objects.get(id=staff_id)
        company_instance = company.objects.get(id=staff.company.id)
        
        if SalesInvoice.objects.count() == 0:
            return render(request, 'sales_invoice_home.html', {"staff": staff})
        else:
            invoices = SalesInvoice.objects.all()
            return render(request, 'sales_invoice_list.html', {'sales_invoices': invoices, "staff": staff})
    
    # Add a default response if 'staff_id' is not in the session
    return HttpResponse("Invalid session or user not logged in.")
      
      
def sales_invoice_home(request):
 return render(request,'sales_invoice_home.html')

def sales_invoice_list(request):
        if 'staff_id' in request.session:
            staff_id = request.session['staff_id']
        else:
            return redirect('/')

        staff = staff_details.objects.get(id=staff_id) 
        invoices = SalesInvoice.objects.all()
        return render(request,'sales_invoice_list.html',{'sales_invoices': invoices,"staff":staff})


def sales_invoice_add(request):
  if 'staff_id' in request.session:
    if request.session.has_key('staff_id'):
      staff_id = request.session['staff_id']
           
    else:
      return redirect('/')
  staff =  staff_details.objects.get(id=staff_id)
  company_instance = company.objects.get(id=staff.company.id)

  Party = Parties.objects.all()
  item = ItemModel.objects.filter(company=company_instance)
  item_units = ItemUnitModel.objects.filter(company=company_instance)
  cust = Parties.objects.filter(company=company_instance)
  if SalesInvoice.objects.filter(company=company_instance).exists():
        invoice_count = SalesInvoice.objects.last().invoice_no
        next_count = invoice_count+1
  else:
        next_count=1

  return render(request, 'sales_invoice_add.html',{'staff':staff,'Party':Party,'item':item,'count':next_count,'item_units':item_units,'cust':cust})


def salesinvoice_add_parties(request):
    if request.method == 'POST':
        if 'staff_id' in request.session:
            staff_id = request.session['staff_id']
        else:
            return redirect('/')

        staff = staff_details.objects.get(id=staff_id)
        company_instance = staff.company 
        
        party_name = request.POST['partyname']
        contact = request.POST['mobilenumber']
        gstin = request.POST['gstin']
        gst_type = request.POST['gstintype']
        state = request.POST['state']
        email = request.POST['email']
        date = request.POST['date']
        address = request.POST['address']
        opening_balance = request.POST.get('balance', '')
        comp=company_instance
        if (
          not party_name
          
      ):
          return render(request, 'add_salesinvoice.html')

        part = Parties(party_name=party_name, gstin=gstin,phone_number=contact,gst_type=gst_type, state=state,billing_address=address, email=email, opening_balance=opening_balance,
                       date=date,company=comp,staff=staff)
        part.save() 

        if 'edit_invoice' in request.POST:
            return redirect('sales_invoice_list')  # Redirect to the invoice listing page
        else:
            return redirect('sales_invoice_add')  # Redirect to the add new invoice page

    return render(request, 'sales_invoice_add.html')


@csrf_exempt
@require_POST
def salesinvoice_saveitem(request):
  if request.method == 'POST':
    sid = request.session.get('staff_id')
    staff =  staff_details.objects.get(id=sid)
    cmp = company.objects.get(id=staff.company.id)

    name = request.POST['name']
    unit = request.POST['unit']
    hsn = request.POST['hsn']
    taxref = request.POST['taxref']
    sell_price = request.POST['sell_price']
    cost_price = request.POST['cost_price']
    intra_st = request.POST['intra_st']
    inter_st = request.POST['inter_st']

    if taxref != 'Taxable':
        intra_st = 'GST0[0%]'
        inter_st = 'IGST0[0%]'

    itmdate = request.POST.get('itmdate')
    stock = request.POST.get('stock')
    itmprice = request.POST.get('itmprice')
    minstock = request.POST.get('minstock')

    # Check if the HSN already exists
    if ItemModel.objects.filter(item_hsn=hsn,company=cmp).exists():
       messages.info(request, 'Sorry, HSN Number already exists')
       return redirect('add_debitnote')
    else:
        itm = ItemModel(item_name=name, item_hsn=hsn,item_unit=unit,item_taxable=taxref, item_gst=intra_st,item_igst=inter_st, item_sale_price=sell_price, 
                    item_purchase_price=cost_price,item_opening_stock=stock,item_current_stock=stock,item_at_price=itmprice,item_date=itmdate,
                    item_min_stock_maintain=minstock,company=cmp,user=cmp.user)
        itm.save() 
        return JsonResponse({'success': True})
    
  else:
        return JsonResponse({'success': False, 'message': 'Invalid request method'})
      
def salesinvoice_item_dropdown(request):
  sid = request.session.get('staff_id')
  staff =  staff_details.objects.get(id=sid)
  cmp = company.objects.get(id=staff.company.id)
  product = ItemModel.objects.filter(company=cmp,user=cmp.user)

  id_list = []
  product_list = []
  items = {}
  for p in product:
    id_list.append(p.id)
    product_list.append(p.item_name)
    items[p.id] = [p.id, p.item_name, p.item_hsn, p.item_sale_price, p.item_gst,p.item_igst]
  return JsonResponse(items)


def save_sales_invoice(request):
    if request.method == 'POST':
        if 'staff_id' in request.session:
            staff_id = request.session['staff_id']
        else:
            return redirect('/')

        staff = staff_details.objects.get(id=staff_id)
        company_instance = staff.company 

        # Get party_id from the form data
        party_id = request.POST.get("customername")
        
        if party_id:
            # If party_id is provided, try to get the Party instance
            try:
                Party = Parties.objects.get(id=party_id)
                party_name = Party.party_name
            except Parties.DoesNotExist:
                # Handle the case where the party is not found
                return HttpResponse("Party not found", status=400)
        else:
            # If party_id is not provided, set Party and party_name to None
            Party = None
            party_name = None

        contact = request.POST.get('phno')
        address = request.POST.get('bname')
        invoice_no = request.POST.get('invoiceno')
        date = request.POST.get('date')
        state_of_supply = request.POST.get('state_of_supply')
        paymenttype = request.POST.get('bank')
        cheque = request.POST.get('chequeNumber')
        upi = request.POST.get('upiNumber')
        accountno = request.POST.get('accountNumber')
        product = request.POST.getlist("product[]")
        hsn = request.POST.getlist("hsn[]")
        qty = request.POST.getlist("qty[]")
        rate = request.POST.getlist("price[]")
        discount = request.POST.getlist("discount[]")
        tax = request.POST.getlist("tax[]")
        total = request.POST.getlist("total[]")
        description = request.POST.get('description')
        advance = request.POST.get("advance")
        balance = request.POST.get("balance")
        subtotal = float(request.POST.get('subtotal'))
        igst = request.POST.get('igst')
        cgst = request.POST.get('cgst')
        sgst = request.POST.get('sgst')
        adjust = request.POST.get("adj")
        taxamount = request.POST.get("taxamount")
        grandtotal = request.POST.get('grandtotal')

        # Create a SalesInvoice instance
        sales_invoice = SalesInvoice(
            staff=staff,
            company=company_instance,
            party_name=party_name,
            party=Party,
            contact=contact,
            address=address,
            invoice_no=invoice_no,
            date=date,
            state_of_supply=state_of_supply,
            paymenttype=paymenttype,
            cheque=cheque,
            upi=upi,
            accountno=accountno,
            description=description,
            subtotal=subtotal,
            igst=igst,
            cgst=cgst,
            sgst=sgst,
            total_taxamount=taxamount,
            adjustment=adjust,
            grandtotal=grandtotal,
            paidoff=advance,
            totalbalance=balance,
            action="CREATED"
        )

        # Save the SalesInvoice instance
        sales_invoice.save()

        # Create a SalesInvoice_History instance
        transaction_history = SalesInvoice_History(
            company=company_instance,
            staff=staff,
            salesinvoice=sales_invoice,
            action="CREATED",
        )
        transaction_history.save()

        # Retrieve the created SalesInvoice instance
        invoice = SalesInvoice.objects.get(id=sales_invoice.id)

        # Iterate through the indices of one of the lists
        for i in range(len(product)):
            product_id = int(product[i])
            itm = ItemModel.objects.get(id=product_id)

            # Create SalesInvoiceItem instances
            SalesInvoiceItem.objects.create(
                item=itm,
                hsn=hsn[i] if i < len(hsn) else None,
                quantity=qty[i] if i < len(qty) else None,
                rate=rate[i] if i < len(rate) else None,
                discount=discount[i] if i < len(discount) else None,
                tax=tax[i] if i < len(tax) else None,
                totalamount=total[i] if i < len(total) else None,
                salesinvoice=invoice,
                company=company_instance,
                staff=staff
            )

        if 'save_and_new' in request.POST:
            return redirect('sales_invoice_add')
        else:
            return redirect('sales_invoice_list')

    return render(request, 'sales_invoice_add.html')
  
  
def edit_salesinvoice(request,id):
  if 'staff_id' in request.session:
    if request.session.has_key('staff_id'):
      staff_id = request.session['staff_id']
           
    else:
      return redirect('/')
  staff =  staff_details.objects.get(id=staff_id)
  company_instance = company.objects.get(id=staff.company.id)
  getinoice=SalesInvoice.objects.get(id=id,company=company_instance)
  getitem=SalesInvoiceItem.objects.filter(salesinvoice=id,company=company_instance)
  Party=Parties.objects.filter(company=company_instance)
  item=ItemModel.objects.filter(company=company_instance)
  

  return render(request, 'edit_salesinvoice.html',{'staff':staff,'getinoice':getinoice,'getitem':getitem,'Party':Party,'item':item})
  

def edit_save_sales_invoice(request,id):

    if request.method == 'POST':
        
        if 'staff_id' in request.session:
            staff_id = request.session['staff_id']
        else:
            return redirect('/')

        staff = staff_details.objects.get(id=staff_id)
        company_instance = staff.company 
        
    
        sales_invoice=SalesInvoice.objects.get(id=id)
        
        sales_invoice.party_name = request.POST.get('partyname')
        sales_invoice.contact = request.POST.get('contact')
        sales_invoice.address = request.POST.get('address')
        sales_invoice.invoice_no = request.POST.get('invoiceno')
        sales_invoice.date = request.POST.get('date')
        sales_invoice.state_of_supply = request.POST.get('state_of_supply')
        sales_invoice.paymenttype = request.POST.get('bank')
        sales_invoice.cheque = request.POST.get('chequeNumber')
        sales_invoice.upi = request.POST.get('upiNumber')
        sales_invoice.accountno = request.POST.get('accountNumber')
        sales_invoice.product = request.POST.getlist("product[]")
        sales_invoice.hsn =  request.POST.getlist("hsn[]")
        sales_invoice.qty =  request.POST.getlist("qty[]")
        sales_invoice.rate = request.POST.getlist("price[]")
        sales_invoice.discount =  request.POST.getlist("discount[]")
        sales_invoice.tax =  request.POST.getlist("tax")
        sales_invoice.total = request.POST.getlist("total[]")
        sales_invoice.description = request.POST.get('description')
        sales_invoice.advance = request.POST.get("advance")
        sales_invoice.balance = request.POST.get("balance")
        sales_invoice.subtotal = float(request.POST.get('subtotal'))
        sales_invoice.igst = request.POST.get('igst')
        sales_invoice.cgst = request.POST.get('cgst')
        sales_invoice.sgst = request.POST.get('sgst')
        sales_invoice.adjust = request.POST.get("adj")
        sales_invoice.taxamount = request.POST.get("taxamount")
        sales_invoice.grandtotal = request.POST.get('grandtotal')
        sales_invoice.action="UPDATED"
        sales_invoice.save()
        try:
            party = Parties.objects.get(party_name=sales_invoice.party_name)
        except Parties.DoesNotExist:
            # Handle the case where the party is not found
            party = None  # or any default value you want to assign
        
      
       
        product = tuple(request.POST.getlist("product[]"))
        qty = tuple(request.POST.getlist("qty[]"))
        tax =tuple( request.POST.getlist("tax[]"))
        discount = tuple(request.POST.getlist("discount[]"))
        total = tuple(request.POST.getlist("total[]"))
        SalesInvoiceItem.objects.filter(salesinvoice=sales_invoice,company=company_instance).delete()
        if len(product)==len(qty)==len(qty)==len(tax):
          mapped=zip(product,qty,tax,discount,total)
          mapped=list(mapped)
          for ele in mapped:
            itm = ItemModel.objects.get(id=ele[0])
            SalesInvoiceItem.objects.create(item =itm,quantity=ele[1], tax=ele[2],discount=ele[3],totalamount=ele[4],salesinvoice=sales_invoice,company=company_instance)

        tr_history = SalesInvoice_History(company=company_instance,
                                              staff=staff,      
                                              salesinvoice=sales_invoice,
                                              action="UPDATED",
                                              
                                              )
        tr_history.save()
        
        
        return redirect('sales_invoice_list')



    return render(request, 'sales_invoice_add.html')


def delete_sales_invoice(request,id):
    salesinvoice=SalesInvoice.objects.get(id=id)
    salesinvoice_item = SalesInvoiceItem.objects.filter(salesinvoice=salesinvoice)
    salesinvoice_history=SalesInvoice_History.objects.filter(salesinvoice=salesinvoice)
    salesinvoice.delete()
    salesinvoice_item.delete()
    salesinvoice_history.delete()
    return redirect('sales_invoice_list')
  
def salesinvoice_history(request,id):
  if 'staff_id' in request.session:
    if request.session.has_key('staff_id'):
      staff_id = request.session['staff_id']
           
    else:
      return redirect('/')
    staff =  staff_details.objects.get(id=staff_id)
    company_instance = company.objects.get(id=staff.company.id)
  
  
  history= SalesInvoice_History.objects.filter(salesinvoice=id)
  return render(request, 'salesinvoice_history.html',{'history':history,"staff":staff})


def itemdata_salesinvoiceedit(request):
  itmid = request.GET['id']
  print(itmid)
  itm = ItemUnitModel.objects.get(id=itmid)
  print(itm)
  hsn = itm.item_hsn
  gst = itm.item_gst
  igst = itm.item_igst
  price = itm.item_sale_price
  qty = itm.item_current_stock
  return JsonResponse({'hsn':hsn, 'gst':gst, 'igst':igst, 'price':price, 'qty':qty})


def salesinvoice_template(request,id):
  if 'staff_id' in request.session:
    if request.session.has_key('staff_id'):
      staff_id = request.session['staff_id']
           
    else:
      return redirect('/')
  staff =  staff_details.objects.get(id=staff_id)
  history= SalesInvoice_History.objects.filter(salesinvoice=id)
  salesinvoice = SalesInvoice.objects.get(id=id)
  salesinvoiceitem = SalesInvoiceItem.objects.filter(salesinvoice=salesinvoice)
  dis = 0
  for itm in salesinvoiceitem:
    dis += int(itm.discount)
  itm_len = len(salesinvoiceitem)
  return render(request, 'salesinvoice_template.html',{'staff':staff,'history':history,'salesinvoice':salesinvoice,'salesinvoiceitem':salesinvoiceitem,'dis':dis,'itm_len':itm_len})


def salesinvoice_graph(request):
    if 'staff_id' in request.session:
        staff_id = request.session['staff_id']
    else:
        return redirect('/')

    staff = staff_details.objects.get(id=staff_id)
    company_instance = company.objects.get(id=staff.company.id)
    sales_data = (
        SalesInvoice.objects
        .values('date__month')
        .annotate(total_sales=Sum('grandtotal'))
    )

    # Extract month-wise sales and labels with month names
    month_names = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun', 'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']
    labels = [month_names[month - 1] for month in range(1, 13)]
    sales = [sales_dict.get('total_sales', 0) for sales_dict in sales_data]

    # Prepare data for chart
    chart_data = {'labels': labels, 'sales': sales}
    years = list(range(2022, 2031))
    return render(request, 'salesinvoice_graph.html', {'chart_data': chart_data,'staff':staff,'years':years})


def export_sales_invoices_to_excel(request):
    # Assuming you have a queryset of sales invoices
    sales_invoices = SalesInvoice.objects.all()

    # Create a new Excel workbook and add a worksheet
    wb = Workbook()
    ws = wb.active

    # Define a custom date style for the cells
    date_style = NamedStyle(name='date_style', number_format='YYYY-MM-DD')
    wb.add_named_style(date_style)

    # Write header row
    headers = ['Invoice No', 'Party Name', 'Date', 'Total Amount', 'Transaction', 'Payment', 'Amount', 'Balance', 'Action', 'By']
    for col_num, header in enumerate(headers, 1):
        ws.cell(row=1, column=col_num, value=header)

    # Write data rows
    for row_num, invoice in enumerate(sales_invoices, 2):
        ws.cell(row=row_num, column=1, value=invoice.invoice_no)
        ws.cell(row=row_num, column=2, value=invoice.party_name)

        # Format the date and write it to the cell
        ws.cell(row=row_num, column=3, value=invoice.date.strftime('%Y-%m-%d'))
        ws.cell(row=row_num, column=3).style = 'date_style'

        ws.cell(row=row_num, column=4, value=invoice.grandtotal)
        ws.cell(row=row_num, column=5, value='SALE')
        ws.cell(row=row_num, column=6, value=invoice.paymenttype)
        ws.cell(row=row_num, column=7, value=invoice.grandtotal)
        ws.cell(row=row_num, column=8, value=invoice.totalbalance)
        ws.cell(row=row_num, column=9, value=invoice.action)
        ws.cell(row=row_num, column=10, value=invoice.company.company_name)

    # Create a response with the Excel file
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=sales_invoices.xlsx'
    wb.save(response)

    return response

  
def import_invoice_data(request):
    if request.method == 'POST':
      
      
        if 'staff_id' in request.session:
            staff_id = request.session['staff_id']
        else:
            return redirect('/')

        staff = staff_details.objects.get(id=staff_id)
        company_instance = staff.company 
        excel_file = request.FILES['file']
        workbook = load_workbook(excel_file)
        worksheet = workbook.active

        # Iterate through each row (skip header row)
        for row in worksheet.iter_rows(min_row=2):
            # Extract data from each cell
            invoice_no = row[0].value
            date = row[1].value
            party_name = row[2].value
            Supply_State = row[3].value
            Payment_Method = row[4].value
            Amount = row[5].value
            Adjustment = row[6].value
            Tax_Amount = row[7].value
            ACTION = row[8].value
            Balance = row[9].value
           
            sale_invoice = SalesInvoice.objects.create(
                invoice_no=invoice_no,
                date=date,
                party_name=party_name,
                state_of_supply=Supply_State,
                paymenttype=Payment_Method,
                grandtotal=Amount,
                adjustment=Adjustment,
                total_taxamount=Tax_Amount,
                action=ACTION,
                paidoff=Balance,
                
            )

            
            SalesInvoice_History.objects.create(
            company=company_instance,
            staff=staff,      
            salesinvoice=sale_invoice,
            action="CREATED",
            )

        return redirect('sales_invoice_list')  

    return render(request, 'sales_invoice_list')  
    
    
def import_excel(request):
    print("hello")
    if request.method == "POST" and request.FILES.get("file"):
      staff_id = request.session['staff_id']
      staff =  staff_details.objects.get(id=staff_id)
      print("open============================================")
      excel_file = request.FILES['file']
      if excel_file.name.endswith('.xlsx'):
        print("open1111111111111111111111")
        df = pd.read_excel(excel_file, engine='openpyxl')
        for index, row in df.iterrows():
          print(row['Party Name'])
          s = SalesInvoice(
                    party_name=row['Party Name'],
                    invoice_no=row['Invoice No'],
                    date=row['Date'],
                    state_of_supply=row['Supply State'],
                    grandtotal=row['Amount'],
                    paidoff=row['Balance'],
                    total_taxamount=row['Tax Amount'],
                    action=row['ACTION'],
                    paymenttype=row['Payment Method'],
                    adjustment=row['Adjustment'],
                    totalbalance=row['Balance'],
                    staff=staff,
                    company=staff.company,
                    
                )
          s.save()
          tran= SalesInvoice_History(
            salesinvoice=s,staff=staff,company=staff.company,action="CREATED",date=date.today()
            )
          tran.save()
        print("success============================================")
        return redirect('sales_invoice_list') 
      print("end===========================")
    return redirect('sales_invoice_list')


def shareinvoiceToEmail(request,id):
  if request.method == 'POST':
    print("email")
    emails_string = request.POST['email_ids']
                # Split the string by commas and remove any leading or trailing whitespace
    emails_list = [email.strip() for email in emails_string.split(',')]
    email_message = request.POST['email_message']
    print(emails_list)

    sid = request.session.get('staff_id')
    staff =  staff_details.objects.get(id=sid)
    cmp = company.objects.get(id=staff.company.id) 
    sale = SalesInvoice.objects.get(id=id,staff=staff)
    saleitem= SalesInvoiceItem.objects.filter(salesinvoice=sale.id)
    context = {'sale':sale, 'cmp':cmp,'saleitem':saleitem}
    template_path = 'salesinvoice_email.html'
    salesinvoice = SalesInvoice.objects.get(id=id,company=cmp)
    salesinvoiceitem = SalesInvoiceItem.objects.filter(salesinvoice=salesinvoice,company=cmp)
                        
    context = {'salesinvoice':salesinvoice, 'cmp':cmp,'salesinvoiceitem':salesinvoiceitem}
    template = get_template(template_path)
    html  = template.render(context)
    result = BytesIO()
    pdf = pisa.pisaDocument(BytesIO(html.encode("ISO-8859-1")), result)
    pdf = result.getvalue()
    filename = f'SALESINVOICE - {sale.invoice_no}.pdf'
    subject = f"SALESINVOICE - {sale.invoice_no}"
    email = EmailMessage(subject, f"Hi,\nPlease find the attached SALESINVOICE - File-{sale.invoice_no}. \n{email_message}\n\n--\nRegards,\n{cmp.company_name}\n{cmp.address}\n{cmp.state} - {cmp.country}\n{cmp.contact}", from_email=settings.EMAIL_HOST_USER, to=emails_list)
    email.attach(filename, pdf, "application/pdf")
    email.send(fail_silently=False)
    # msg = messages.success(request, 'Debit note file has been shared via email successfully..!')
    return redirect(salesinvoice_template, id)
  
  
def salesinvoice_savecustomer(request):
    print("party1")
    sid = request.session.get('staff_id')
    staff = staff_details.objects.get(id=sid)
    cmp = company.objects.get(id=staff.company.id)

    party_name = request.POST['name']
    email = request.POST['email']
    
    phone_number = request.POST.get('mobile', None)
    state = request.POST['splystate']
    address = request.POST['baddress']
    gst_type = request.POST['gsttype']
    gst_no = request.POST['gstin']
    current_date = request.POST['partydate']
    openingbalance = request.POST.get('openbalance')
    payment = request.POST.get('paytype')
    
    
    part = Parties(
        party_name=party_name,
        gstin=gst_no,
        phone_number=phone_number,
        gst_type=gst_type,
        state=state,
        billing_address=address,
        email=email,
        opening_balance=openingbalance,
        to_pay=(payment == 'to_pay'),  # Set to True if payment is 'to_pay'
        to_recieve=(payment == 'to_receive'),
        date=current_date,
        company=cmp,
        staff=staff,
        
    )
    
    
    part.save()
    history=History(party=part,staff=staff,company=cmp,action="CREATED",date=current_date)
    history.save()
    
    partyobj = Parties.objects.filter(company=cmp).values('id', 'party_name')

    party_list = [{'id': customer['id'], 'name': customer['party_name']} for customer in partyobj]
    
    return JsonResponse({'party_list':party_list, 'success': True, 'message': 'Customer saved successfully.'})
  
  
def salesinvoice_add_unit(request):
    if request.method == 'POST':
        new_unit_name = request.POST.get('newUnit')

        if new_unit_name:
           
            new_unit_name = new_unit_name.upper()

            
            if not ItemUnitModel.objects.filter(unit_name=new_unit_name).exists():
                
                new_unit = ItemUnitModel(unit_name=new_unit_name)
                new_unit.save()

                messages.success(request, 'Unit added successfully')
                return JsonResponse({'status': 'success', 'message': 'Unit added successfully'})
            else:
                messages.error(request, 'Unit already exists')
                return JsonResponse({'status': 'error', 'message': 'Unit already exists'})
        else:
            messages.error(request, 'Unit name cannot be empty')
            return JsonResponse({'status': 'error', 'message': 'Unit name cannot be empty'})

    return redirect('sales_invoice_add')
    
def check_unit_existence(request):
    newUnit = request.GET.get('newUnit', '').upper()
    exists = ItemUnitModel.objects.filter(unit_name=newUnit).exists()
    return JsonResponse({'exists': exists})

def check_unit_existence1(request):
    newUnit = request.GET.get('newUnit', '').upper()
    exists = ItemUnitModel.objects.filter(unit_name=newUnit).exists()
    return JsonResponse({'exists': exists})
    
def get_all_units(request):
    try:
        # Fetch all units from ItemUnitModel
        units = ItemUnitModel.objects.values_list('unit_name', flat=True)
        
        # Convert the QuerySet to a list
        units_list = list(units)
        
        # Return the list as JSON response
        return JsonResponse({'units': units_list})
    except Exception as e:
        # Handle exceptions if any
        return JsonResponse({'error': str(e)})
    
def get_all_unit1(request):
    try:
        sid = request.session.get('staff_id')
        staff = staff_details.objects.get(id=sid)
        cmp = company.objects.get(id=staff.company.id)
        itmunits = ItemUnitModel.objects.filter(company=cmp)
        
        # Fetch all units from ItemUnitModel
        units = itmunits.values_list('unit_name', flat=True)
        
        # Convert the QuerySet to a list
        units_list = list(units)
        
        # Return the list as JSON response
        return JsonResponse({'units': units_list})
    except Exception as e:
        # Handle exceptions if any
        return JsonResponse({'error': str(e)})
        
def salesinvoice_custdata(request):
    cid = request.GET.get('customer_id')
    part = party.objects.get(id=cid)
    phno = part.phone_number
    address = part.billing_address
    pay = part.payment
    bal = part.openingbalance
    return JsonResponse({'phno': phno, 'address': address, 'pay': pay,'bal':bal})
    
    
# def check_gst_exists(request):
#     gstin= request.GET.get('gstin')
    
#     if party.objects.filter(gstin=gstin).exists():
#         return JsonResponse({'exists': True})
#     return JsonResponse({'exists':False})

def check_gst_exists(request):
    gstin = request.GET.get('gstin')
    sid = request.session.get('staff_id')
    staff = staff_details.objects.get(id=sid)
    cmp = company.objects.get(id=staff.company.id)
    gsttype = request.GET.get('gsttype')  
    print(gsttype)

    if gsttype != "Unregistered or Consumer":
        if Parties.objects.filter(company=cmp, gstin=gstin).exists():
            return JsonResponse({'exists': True})
        return JsonResponse({'exists': False})
    
    
def history_debitnote(request,id):
  sid = request.session.get('staff_id')
  staff =  staff_details.objects.get(id=sid)
  cmp = company.objects.get(id=staff.company.id)  
  allmodules= modules_list.objects.filter(company=staff.company,status='New')
  pdebt = purchasedebit.objects.get(pdebitid=id,company=cmp)
  hsty= DebitnoteTransactionHistory.objects.filter(debitnote=id,company=cmp)
  context = {'staff':staff,'allmodules':allmodules,'hsty':hsty,'id':id}
  return render(request,'debitnotehistory.html',context)

  
def debthistory(request):
    if request.method == 'POST' and 'id' in request.POST:
        pid = request.POST['id']
        sid = request.session.get('staff_id')
        
        if sid:
            staff = get_object_or_404(staff_details, id=sid)
            cmp = get_object_or_404(company, id=staff.company.id)
            pdebt = get_object_or_404(purchasedebit, pdebitid=pid, company=cmp)
            hsty = DebitnoteTransactionHistory.objects.filter(debitnote=pdebt, company=cmp).last()
            
            if hsty:
                name = hsty.staff.first_name + ' ' + hsty.staff.last_name 
                action = hsty.action
                return JsonResponse({'name': name, 'action': action, 'pid': pid})
            else:
                return JsonResponse({'error': 'No transaction history found for this purchase debit.'}, status=404)
        else:
            return JsonResponse({'error': 'Staff ID not found in session.'}, status=404)
    else:
        return JsonResponse({'error': 'Invalid request method or missing ID parameter.'}, status=400)
  
  
def credit_add(request):
  sid = request.session.get('staff_id')
  staff = staff_details.objects.get(id=sid)
  cmp = company.objects.get(id=staff.company.id)
  todaydate = date.today().isoformat()
  party = Parties.objects.filter(company_id=cmp.id)
  item=ItemModel.objects.filter(company_id=cmp.id) 
  last_ref = Creditnote.objects.filter(company=cmp).order_by('-returnno').first()
  if last_ref:
        refno = last_ref.returnno + 1
  else:
        refno = 1
  context = {
        'refno': refno,
        'party':party,
        'item':item,
        'todaydate':todaydate,
    }
  return render(request, 'credit_add.html',context)

def creditbilldata(request):
    try:
        partyid = request.POST['id']
        party_instance = Parties.objects.get(id=partyid)

        # Initialize lists to store multiple bill numbers and dates
        bill_numbers = []
        bill_dates = []

        try:
            # Retrieve all instances for the party
            bill_instances = SalesInvoice.objects.filter(party=party_instance)

            # Loop through each instance and collect bill numbers and dates
            for bill_instance in bill_instances:
                bill_numbers.append(bill_instance.invoice_no)
                bill_dates.append(bill_instance.date)

        except SalesInvoice.DoesNotExist:
            pass

        # Return a JSON response with the list of bill numbers and dates
        if not bill_numbers and not bill_dates:
            return JsonResponse({'bill_numbers': ['nobill'], 'bill_dates': ['nodate']})

        return JsonResponse({'bill_numbers': bill_numbers, 'bill_dates': bill_dates})

    except KeyError:
        return JsonResponse({'error': 'The key "id" is missing in the POST request.'})

    except Parties.DoesNotExist:
        return JsonResponse({'error': 'Party not found.'})
    
def credit_bill_date(request):
    selected_bill_no = request.POST.get('billNo', None)
    print(selected_bill_no)

    try:

        # Get the latest SalesInvoice with the specified bill_number and party
        bill = SalesInvoice.objects.filter(invoice_no=selected_bill_no).latest('date')
        bill_date = bill.date.strftime('%Y-%m-%d')
        print(bill_date)
        
    except SalesInvoice.DoesNotExist:
        return JsonResponse({'error': 'Bill number not found'}, status=400)
    except SalesInvoice.MultipleObjectsReturned:
        return JsonResponse({'error': 'Multiple SalesInvoices found for the same bill number'}, status=400)

    return JsonResponse({'bill_date': bill_date})

def saveparty(request):
    if request.method == 'POST':
        # Retrieve session and company information
        sid = request.session.get('staff_id')
        staff = staff_details.objects.get(id=sid)
        company_obj = company.objects.get(id=staff.company.id)

        # Retrieve party data from POST request
        party_name = request.POST.get('partyname', '').capitalize()
        mobile_number = request.POST.get('mobilenumber', '')
        gstin = request.POST.get('gstin', '')
        gstintype = request.POST.get('gstintype', '')
        state = request.POST.get('state', '')
        email = request.POST.get('email', '')
        date = request.POST.get('date', '')
        address = request.POST.get('address', '')
        balance = request.POST.get('balance', '')

        # Check if party already exists
        if Parties.objects.filter(phone_number=mobile_number, company=company_obj).exists():
            return JsonResponse({'success': False, 'message': 'Contact number already exists'})
        elif Parties.objects.filter(email=email, company=company_obj).exists():
            return JsonResponse({'success': False, 'message': 'Email already exists'})
        elif Parties.objects.filter(gstin=gstin, company=company_obj).exists():
            return JsonResponse({'success': False, 'message': 'GST number already exists'})
        else:
            # Create and save the party
            party = Parties(party_name=party_name, gstin=gstin, phone_number=mobile_number, gst_type=gstintype,
                            state=state, email=email, opening_balance=balance, date=date,billing_address=address, company=company_obj,
                            staff=staff)
            party.save()
            return JsonResponse({'success': True, 'id': party.id})
    else:
        return JsonResponse({'success': False, 'message': 'Invalid request method'})

def party_dropdown(request):
    sid = request.session.get('staff_id')
    staff = staff_details.objects.get(id=sid)
    company_obj = company.objects.get(id=staff.company.id)
    parties = Parties.objects.filter(company=company_obj)

    id_list = [party.id for party in parties]
    party_list = [party.party_name for party in parties]

    return JsonResponse({'id_list': id_list, 'party_list': party_list}) 


def savecredititem(request):
    if request.method == 'POST':
        sid = request.session.get('staff_id')
        staff =  staff_details.objects.get(id=sid)
        cmp = company.objects.get(id=staff.company.id)
        item_name = request.POST.get('item_name')
        hsn = request.POST.get('hsn')
        qty = request.POST.get('qty')
        tax_ref = request.POST.get('taxref')
        intra_st = request.POST.get('intra_st')
        inter_st = request.POST.get('inter_st')
        sale_price = request.POST.get('sale_price')  # Changed to match the AJAX data key
        purchase_price = request.POST.get('purchase_price')  # Changed to match the AJAX data key

        # Check if the HSN already exists
        if ItemModel.objects.filter(item_hsn=hsn).exists():
            return JsonResponse({'success': False, 'message': 'HSN Number already exists'})
        else:
            # Save new item
            item = ItemModel(
                item_name=item_name,
                item_hsn=hsn,
                item_current_stock=qty,
                item_taxable=tax_ref,
                item_gst=intra_st,
                item_igst=inter_st,
                item_sale_price=sale_price,
                item_purchase_price=purchase_price,
                staff=staff,
                company=cmp
                # staff and company fields should be properly set based on your requirements
            )
            item.save()
            return JsonResponse({'success': True})
    else:
        return JsonResponse({'success': False, 'message': 'Invalid request method'})

def credititem_dropdown(request):
  sid = request.session.get('staff_id')
  staff =  staff_details.objects.get(id=sid)
  cmp = company.objects.get(id=staff.company.id)
  print(sid, staff, cmp)
  options = {}
  option_objects = ItemModel.objects.filter(company=cmp)
  for option in option_objects:
      options[option.id] = [option.id, option.item_name]
  return JsonResponse(options)

def credit_save(request):
    if request.method == 'POST':
        sid = request.session.get('staff_id')
        staff = staff_details.objects.get(id=sid)
        cmp = company.objects.get(id=staff.company.id)

        # Retrieve party details if available
        party_id = request.POST.get('partyname')
        party = None
        if party_id:
            party = Parties.objects.get(id=party_id)

        # Check if 'billNo' exists in POST data and assign value accordingly
        bill_no = request.POST.get('billNo', None)
        if bill_no == 'nobill':
            bill_no = None

        # Check if 'billDate' exists in POST data and assign value accordingly
        bill_date_str = request.POST.get('billDate', None)
        bill_date = None
        if bill_date_str:
            bill_date = parse_date(bill_date_str)

        # Create an instance of Creditnote model and save the data
        credit_note = Creditnote(
            party_name=party.party_name if party else None,
            contact=party.phone_number if party else None,
            address=party.billing_address if party else None,
            invoice_no=bill_no,
            idate=bill_date,
            state_of_supply=request.POST.get('placosupply'),
            date=request.POST.get('date'),
            gstin=request.POST.get('gstin'),
            subtotal=request.POST.get('subtotal'),
            sgst=request.POST.get('sgst'),
            cgst=request.POST.get('cgst'),
            igst=request.POST.get('igst'),
            taxamount=request.POST.get('taxamount'),
            roundoff=request.POST.get('adj'),
            grandtotal=request.POST.get('grandtotal'),
            description=request.POST.get('des'),
            returnno=request.POST.get('returnno'),
            staff=staff,
            company=cmp,
            party=party
        )

        # Save the instance
        credit_note.save()
        history = CreditnoteHistory(company=cmp, staff=staff, credit=credit_note, action='Created')
        history.save()

        # Save credit note items
        product = request.POST.getlist("product[]")
        qty = request.POST.getlist("qty[]")
        discount = request.POST.getlist("discount[]")
        total = request.POST.getlist("total[]")
        hsn = request.POST.getlist("hsn[]")
        tax = request.POST.getlist("tax[]")
        price = request.POST.getlist("price[]")

        if len(product) == len(qty) == len(discount) == len(total) == len(hsn) == len(tax) == len(price):
            mapped = zip(product, qty, discount, total, hsn, tax, price)
            for ele in mapped:
                itm = ItemModel.objects.get(id=ele[0])
                CreditnoteItem.objects.create(
                    product=itm.item_name,
                    qty=ele[1],
                    discount=ele[2],
                    total=ele[3],
                    hsn=ele[4],
                    tax=ele[5],
                    price=ele[6],
                    company=cmp,
                    credit=credit_note,
                    staff=staff
                )

        Creditnote.objects.filter(company=cmp, staff=staff).update(returnno=F('returnno'))

        if 'Next' in request.POST:
            return redirect('transactiontable')

        if "Save" in request.POST:
            return redirect('credit_add')

    else:
        return render(request, 'credit_add.html')



def get_tax_rate(request):
    if request.method == 'GET':
        product_id = request.GET.get('id')
        credit_note_id = request.GET.get('credit_note_id')

        if product_id and credit_note_id:
            item = get_object_or_404(ItemModel, id=product_id)
            credit_note = get_object_or_404(Creditnote, id=credit_note_id)

            # Assuming you have a field named place_of_supply in your CreditNote model
            place_of_supply = credit_note.place_of_supply

            # Set tax_rate based on place of supply
            if place_of_supply == 'State':
                tax_rate = item.item_gst
            else:
                tax_rate = item.item_igst

            return JsonResponse({'tax_rate': tax_rate})
    
    return JsonResponse({'error': 'Invalid request'})


def transactiontable(request):
  sid = request.session.get('staff_id')
  staff =  staff_details.objects.get(id=sid)
  cmp = company.objects.get(id=staff.company.id)
  credit=Creditnote.objects.filter(company=cmp)
  item=Creditnote.objects.filter(company_id=cmp.id)
  return render(request,'transaction_table.html',{'credit':credit,'item':item})

def edit_credit(request,pk):
  toda = date.today()
  tod = toda.strftime("%Y-%m-%d")
  sid = request.session.get('staff_id')
  staff =  staff_details.objects.get(id=sid)
  cmp = company.objects.get(id=staff.company.id)
  party = Parties.objects.filter(company=cmp,staff=staff)
  item = ItemModel.objects.filter(company=cmp,staff=staff)
  crd = Creditnote.objects.get(id=pk,company=cmp)
  crditem = CreditnoteItem.objects.filter(credit=crd,company=cmp)
  cdate = crd.date.strftime("%Y-%m-%d")
  print(crd.idate)
  context = {
    'staff':staff,  
    'crd':crd, 
    'crditem':crditem, 
    'party':party, 
    'item':item,
    'tod':tod,
    'cdate':cdate
    }
  
  return render(request,'creditnote_edit.html',context)

def update_creditnote(request, pk):
    if request.method == 'POST':
        sid = request.session.get('staff_id') 
        staff = get_object_or_404(staff_details, id=sid)
        cmp = get_object_or_404(company, id=staff.company.id)

        # Retrieve party details if available
        party_id = request.POST.get('partyname')
        party = get_object_or_404(Parties, id=party_id) if party_id else None

        crd = get_object_or_404(Creditnote, id=pk, company=cmp)
        crd.party = party
        crd.contact = party.phone_number if party else None
        crd.address = party.billing_address if party else None
        crd.date = parse_date(request.POST.get('date1', None))
        crd.invoice_no = request.POST.get('billNo')
        crd.idate = parse_date(request.POST.get('billDate', None))
        crd.state_of_supply = request.POST.get('placosupply')
        crd.subtotal = float(request.POST.get('subtotal', 0))
        crd.grandtotal = request.POST.get('grandtotal')
        crd.igst = request.POST.get('igst')
        crd.cgst = request.POST.get('cgst')
        crd.sgst = request.POST.get('sgst')
        crd.taxamount = request.POST.get("taxamount")
        crd.roundoff = request.POST.get("adj")
        crd.description = request.POST.get("des")

        crd.save()

        product = tuple(request.POST.getlist("product[]"))
        qty = tuple(request.POST.getlist("qty[]"))
        total = tuple(request.POST.getlist("total[]"))
        discount = tuple(request.POST.getlist("discount[]"))
        hsn = request.POST.getlist("hsn[]")
        tax = request.POST.getlist("tax[]")
        price = request.POST.getlist("price[]")

        CreditnoteItem.objects.filter(credit=crd).delete()
        if len(product) == len(qty) == len(discount) == len(total) == len(hsn) == len(tax) == len(price):
          mapped=zip(product, qty, discount, total, hsn, tax, price)
          mapped=list(mapped)
          for ele in mapped:
            itm = ItemModel.objects.get(id=ele[0])
            CreditnoteItem.objects.create(product =itm.item_name,qty=ele[1],discount=ele[2],total=ele[3],hsn=ele[4],tax=ele[5],price=ele[6],credit=crd,company=cmp,item=itm,staff=staff)

        CreditnoteHistory.objects.create(credit=crd,company=cmp,staff=staff,action='Updated')
        return redirect('transactiontable')

    return redirect('transactiontable')


def template1(request,pk):
  sid = request.session.get('staff_id')
  staff = staff_details.objects.get(id=sid)
  cmp = company.objects.get(id=staff.company.id)   
  cd=Creditnote.objects.get(id=pk)
  crditem = CreditnoteItem.objects.filter(credit=cd,company=cmp)
  return render(request,'creditnote1.html',{'cd':cd,'crditem':crditem})

def template2(request,pk):
  sid = request.session.get('staff_id')
  staff = staff_details.objects.get(id=sid)
  cmp = company.objects.get(id=staff.company.id)  
  cd=Creditnote.objects.get(id=pk)
  crditem = CreditnoteItem.objects.filter(credit=cd,company=cmp)
  return render(request,'creditnote2.html',{'cd':cd,'crditem':crditem})

def template3(request,pk):
  sid = request.session.get('staff_id')
  staff = staff_details.objects.get(id=sid)
  cmp = company.objects.get(id=staff.company.id)  
  cd=Creditnote.objects.get(id=pk)
  crditem = CreditnoteItem.objects.filter(credit=cd,company=cmp)
  return render(request,'creditnote3.html',{'cd':cd,'crditem':crditem})

def credithistory(request,pk):
  sid = request.session.get('staff_id')
  staff = staff_details.objects.get(id=sid)
  cmp = company.objects.get(id=staff.company.id)  
  cd=Creditnote.objects.get(id=pk)
  history=CreditnoteHistory.objects.filter(credit=cd,company=cmp)
  context = {'staff':staff,'history':history,'cd':cd}
  return render(request,'credithistory.html',context)


def delete_credit(request,pk):
  sid = request.session.get('staff_id')
  staff = staff_details.objects.get(id=sid)
  cmp = company.objects.get(id=staff.company.id) 
  crd = Creditnote.objects.get(id=pk)
  CreditnoteItem.objects.filter(credit=crd,company=cmp).delete()
  crd.delete()
  return redirect('transactiontable')

def pdftomailcredit(request,pk):
  if request.user:
        try:
            if request.method == 'POST':
                emails_string = request.POST['email_ids']

                # Split the string by commas and remove any leading or trailing whitespace
                emails_list = [email.strip() for email in emails_string.split(',')]
                email_message = request.POST['email_message']
                print(emails_list)
                sid = request.session.get('staff_id')
                staff = staff_details.objects.get(id=sid)
                cmp = company.objects.get(id=staff.company.id) 
                
                cd = Creditnote.objects.get(id=pk,company=cmp)
                itm = CreditnoteItem.objects.filter(credit=cd,company=cmp)
                context = {'cd':cd, 'cmp':cmp,'itm':itm}
                template_path = 'creditmail.html'
                template = get_template(template_path)

                html  = template.render(context)
                result = BytesIO()
                pdf = pisa.pisaDocument(BytesIO(html.encode("ISO-8859-1")), result)
                pdf = result.getvalue()
                filename = f'CREDIT NOTE - {cd.returnno}.pdf'
                subject = f"CREDIT NOTE - {cd.returnno}"
                email = EmailMessage(subject, f"Hi,\nPlease find the attached CREDIT NOTE - File-{cd.returnno}. \n{email_message}\n\n--\nRegards,\n{cmp.company_name}\n{cmp.address}\n{cmp.state} - {cmp.country}\n{cmp.contact}", from_email=settings.EMAIL_HOST_USER, to=emails_list)
                email.attach(filename, pdf, "application/pdf")
                email.send(fail_silently=False)

                # msg = messages.success(request, 'Credit note file has been shared via email successfully..!')
                return redirect('template1',pk=pk)

        except Exception as e:
            print(e)
            messages.error(request, f'{e}')
            return redirect('template1',pk=pk)
          
def partydata(request):
    if request.method == 'POST':
        cid = request.POST.get('id')
        part = Parties.objects.get(id=cid)
        phno = part.phone_number
        address = part.billing_address
        pay = part.to_pay
        bal = part.opening_balance
        return JsonResponse({'phno': phno, 'address': address, 'pay': pay, 'bal': bal})
    else:
        return JsonResponse({'error': 'Invalid request method'})
        
        
def saveitemc(request):
  if request.method == 'POST':
    sid = request.session.get('staff_id')
    staff = staff_details.objects.get(id=sid)
    company_obj = staff.company

    name = request.POST.get('name')
    print(name)
    unit = request.POST.get('unit')
    hsn = request.POST.get('hsn')
    taxref = request.POST.get('taxref')
    sell_price = request.POST.get('sell_price')
    cost_price = request.POST.get('cost_price')
    intra_st = request.POST.get('intra_st')
    inter_st = request.POST.get('inter_st')
    itmdate = request.POST.get('itmdate')
    stock = request.POST.get('stock')
    itmprice = request.POST.get('itmprice')
    minstock = request.POST.get('minstock')

    # if not hsn:
    #     hsn = None

    # Check if HSN exists for the given company
    if ItemModel.objects.filter(company=company_obj, item_hsn=hsn).exists():
        return JsonResponse({'success': False, 'message': 'HSN already exists'})

    itm = ItemModel(
        item_name=name, item_hsn=hsn, item_unit=unit, item_type='Type', item_taxable=taxref, item_gst=intra_st,
        item_igst=inter_st, item_sale_price=sell_price, item_purchase_price=cost_price, item_current_stock=stock,
        item_stock_in_hand=stock, item_at_price=itmprice, item_date=itmdate, company=company_obj, user=company_obj.user
    )
    itm.save()

    return JsonResponse({'success': True})
  
  return JsonResponse({'success': False, 'message': 'Invalid request method'})


def item_dropdownc(request):
  sid = request.session.get('staff_id')
  staff =  staff_details.objects.get(id=sid)
  cmp = company.objects.get(id=staff.company.id)
  print(sid, staff, cmp)
  options = {}
  option_objects = ItemModel.objects.filter(company=cmp)
  for option in option_objects:
      options[option.id] = [option.id, option.item_name]
  return JsonResponse(options)